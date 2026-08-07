# -*- coding: utf-8 -*-
# Altera Excel 生成工具 - 打包版资源路径注入
import os as _os
import sys as _sys

def _get_resource_path():
    """获取资源根目录。
    - PyInstaller 打包后（frozen）：使用 exe 所在目录
      （data 文件夹放在 exe 同目录，随 exe 一起分发）
    - 源码运行：使用脚本所在目录
    """
    if getattr(_sys, "frozen", False):
        # PyInstaller 打包后：用 exe 所在目录
        return _os.path.dirname(_os.path.abspath(_sys.executable))
    # 源码运行：使用脚本所在目录
    script_dir = _os.path.dirname(_os.path.abspath(__file__))
    return script_dir

RESOURCE_PATH = _get_resource_path()
import csv
import json
import os
import re
import sys
import random
import string
import shutil
import zipfile
import tempfile

try:
    csv.field_size_limit(sys.maxsize)
except (OverflowError, ValueError):
    csv.field_size_limit(2147483647)
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from datetime import datetime

try:
    import urllib.request as urlrequest
    _HAS_URLLIB = True
except ImportError:
    _HAS_URLLIB = False

APP_VERSION = "1.5.8"
UPDATE_CHECK_URL = "https://dingshan985-web.github.io/doujiao/version.json"


def _run_update_worker(extract_dir, app_dir, tmp_dir):
    """独立更新进程：等待主程序退出后，复制文件并重启。"""
    import time
    exe_name = "豆脚AlteraExcel工具.exe"
    try:
        time.sleep(2)
        def copy_tree(src, dst):
            if not os.path.exists(dst):
                os.makedirs(dst)
            for item in os.listdir(src):
                s = os.path.join(src, item)
                d = os.path.join(dst, item)
                if os.path.isdir(s):
                    copy_tree(s, d)
                else:
                    shutil.copy2(s, d)
        for item in os.listdir(extract_dir):
            src = os.path.join(extract_dir, item)
            dst = os.path.join(app_dir, item)
            if os.path.isdir(src):
                copy_tree(src, dst)
            else:
                shutil.copy2(src, dst)
        time.sleep(0.5)
        new_exe = os.path.join(app_dir, exe_name)
        if os.path.exists(new_exe):
            if sys.platform == "win32":
                os.startfile(new_exe)
            else:
                import subprocess
                subprocess.Popen([new_exe])
    except Exception as e:
        try:
            log_path = os.path.join(tempfile.gettempdir(), "doujiao_update_error.log")
            with open(log_path, "w", encoding="utf-8") as f:
                f.write(f"Update failed: {e}\n")
                f.write(f"extract_dir: {extract_dir}\n")
                f.write(f"app_dir: {app_dir}\n")
                f.write(f"tmp_dir: {tmp_dir}\n")
        except Exception:
            pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装依赖: pip install openpyxl")
    sys.exit(1)

# ============================================================
# Content Generation Engine - SEO 标题/描述 + Pages 内容生成
# ============================================================

def _load_templates():
    """加载 templates.json。返回 (categories, title_templates, description_templates, avoid_words, defaults)"""
    script_dir = RESOURCE_PATH
    json_path = os.path.join(script_dir, "data", "templates.json")
    fallback = {
        "categories": {"general": {"keywords": ["products"], "adjectives": ["premium"], "benefits": ["Carefully curated for quality and value"]}},
        "title_templates": ["{brand} Outlet Store | {cat_adj} {cat_kw} at Discount Prices"],
        "description_templates": ["Shop {brand} Outlet for premium {cat_kw} at direct discount prices. {benefit}. Free shipping worldwide."],
        "avoid_words": ["replica", "copy", "fake", "duplicate"],
        "defaults": {"keyword": "products", "adjective": "premium", "benefit": "Carefully curated for quality and value"},
    }
    if not os.path.exists(json_path):
        return fallback["categories"], fallback["title_templates"], fallback["description_templates"], fallback["avoid_words"], fallback["defaults"]
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return (
            data.get("categories", fallback["categories"]),
            data.get("title_templates", fallback["title_templates"]),
            data.get("description_templates", fallback["description_templates"]),
            data.get("avoid_words", fallback["avoid_words"]),
            data.get("defaults", fallback["defaults"]),
        )
    except Exception:
        return fallback["categories"], fallback["title_templates"], fallback["description_templates"], fallback["avoid_words"], fallback["defaults"]


_CATEGORIES, _TITLE_TEMPLATES, _DESC_TEMPLATES, _AVOID_WORDS, _DEFAULTS = _load_templates()


def extract_keywords(text):
    """从原站文本中提取有意义的产品关键词和主题词"""
    if not text:
        return []
    text = text.lower()
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'[^\w\s\-]', ' ', text)
    words = re.findall(r'[a-z]{2,}', text)
    stopwords = set("""
        the a an and or but if then so for of to in on at by from as is are was were be been being
        our we you your i me my this that these those it its they them their with without
        official store shop online best free shipping sale new top brand quality products
        buy now today world famous popular worldwide authentic original genuine browse shopify
    """.split())
    meaningful = [w for w in words if w not in stopwords and len(w) >= 3]
    return meaningful


def detect_category(title, description):
    """根据内容判断产品类别（用于选择合适的描述语言）"""
    combined = (title + " " + description).lower()
    scores = {}
    for cat, data in _CATEGORIES.items():
        score = 0
        for kw in data["keywords"]:
            if kw in combined:
                score += 1
        scores[cat] = score
    return max(scores, key=scores.get) if max(scores.values()) > 0 else None


def extract_brand_name(title):
    """从原站 title 中尝试提取品牌名"""
    if not title:
        return "Our"
    cleaned = re.sub(r'[^\w\s&]', ' ', title)
    words = [w for w in cleaned.split() if w and not w.lower() in {"the", "a", "an", "official", "store", "shop", "online", "free", "shipping"}]
    if words:
        brand = words[0].strip()
        if len(brand) >= 2:
            return brand
    return "Premium"


# ============================================================
# 产品关键词提取（方案一：CSV 产品关键词锁定）
# ============================================================

_STOPWORDS = {
    "the", "a", "an", "and", "or", "for", "with", "by", "in", "on", "at",
    "to", "of", "from", "is", "are", "was", "were", "be", "been", "being",
    "has", "have", "had", "do", "does", "did", "will", "would", "could",
    "should", "may", "might", "must", "can", "this", "that", "these",
    "those", "it", "its", "all", "each", "every", "any", "some", "no",
    "not", "but", "if", "then", "so", "because", "as", "until", "while",
    "about", "after", "before", "above", "below", "between", "into",
    "through", "during", "out", "off", "over", "under", "again", "further",
    "once", "here", "there", "when", "where", "why", "how", "only", "same",
    "new", "sale", "buy", "shop", "store", "official", "online", "free",
    "shipping", "discount", "best", "top", "hot", "popular", "latest",
    "cheap", "price", "prices", "review", "reviews", "upc", "ean",
    "men", "women", "kids", "mens", "womens", "man", "woman", "boy", "girl",
    "with", "without", "inch", "inches", "mm", "cm", "m", "ft", "watt", "watts", "v", "ah", "mah",
    "pack", "set", "piece", "pcs", "pair", "pairs", "box", "bag",
    "size", "sizes", "color", "colors", "colour", "colours", "type", "types",
    "model", "models", "style", "styles", "design", "designs",
    "quality", "high", "low", "good", "great", "nice", "perfect",
    "professional", "grade", "series", "generation", "gen",
    "usb", "cable", "charger", "battery", "power",
    "up", "down", "left", "right", "front", "back",
    "more", "less", "most", "least", "very", "really", "just",
    "than", "too", "also", "well", "much", "many", "few",
    "other", "another", "such", "like", "vs", "versus",
}

_PRODUCT_SUFFIXES = {
    "pro", "max", "mini", "air", "slim", "xl", "xxl", "lg", "sm",
    "plus", "standard", "premium", "lite", "classic", "basic",
    "advanced", "ultra", "essential", "ultimate", "elite", "core",
    "sport", "fit", "active", "original", "edition", "v2", "v3",
    "black", "white", "blue", "red", "green", "grey", "gray",
    "leather", "nylon", "canvas", "aluminum", "steel",
    "smart", "digital",
    "small", "large", "medium",
}


def extract_product_keywords_from_csv(csv_path, max_keywords=30):
    """从 CSV 产品数据中提取核心产品关键词，用于锁定生成内容方向。

    关键词来源及权重（以 Title 为主要来源）：
      1. Title 列（主要来源，权重 ×3，同一标题内去重）
      2. Handle 列（辅助来源，权重 ×1）
      3. Tags 列（辅助来源，权重 ×1）

    过滤规则：
      - 排除 _STOPWORDS（介词/连词/通用词/单位词等）
      - 排除 _PRODUCT_SUFFIXES（pro/max/mini 等型号后缀，只保留品类词）
      - 排除纯阿拉伯数字
      - 只保留 3 个字符以上的词
      - 最终返回去重后的关键词列表（按出现频率排序）

    返回：
      dict: {
        "keywords": ["headphone", "bone conduction", ...],  # 产品核心词
        "product_types": ["headphones", "earbuds", ...],    # 产品类型
        "top_handles": ["open-ear-headphones", ...],        # 高频 handle 词
        "all_text": "full combined text for prompt use",    # 拼接全文供 AI 参考
      }
    """
    import collections

    if not csv_path or not os.path.exists(csv_path):
        return None

    try:
        headers, rows = read_product_csv(csv_path)
    except Exception:
        return None

    if not headers or not rows:
        return None

    # 找到关键列索引（不区分大小写匹配）
    def col_index(name):
        for i, h in enumerate(headers):
            if h.strip().lower() == name.strip().lower():
                return i
        return -1

    title_idx = col_index("title")
    handle_idx = col_index("handle")
    tags_idx = col_index("tags")

    word_freq = collections.Counter()
    product_types = set()
    all_words = []

    for row in rows:
        # 从 Title 列提取（主要来源，权重最高）
        title_part = row[title_idx].lower() if title_idx >= 0 and title_idx < len(row) else ""
        title_words = re.findall(r'[a-z][a-z0-9]*', title_part)
        seen_title = set()
        for w in title_words:
            if len(w) < 3:
                continue
            if w.isdigit():
                continue
            if w in _STOPWORDS:
                continue
            if w in _PRODUCT_SUFFIXES:
                continue
            if w in seen_title:
                continue
            seen_title.add(w)
            word_freq[w] += 3
            all_words.append(w)

        # 从 Handle 列提取（辅助来源，权重较低）
        handle_part = row[handle_idx].lower() if handle_idx >= 0 and handle_idx < len(row) else ""
        handle_words = re.findall(r'[a-z][a-z0-9]*', handle_part)
        for w in handle_words:
            if len(w) < 3:
                continue
            if w.isdigit():
                continue
            if w in _STOPWORDS:
                continue
            if w in _PRODUCT_SUFFIXES:
                continue
            word_freq[w] += 1

        # 从 Tags 列提取（辅助来源，权重较低）
        tags_part = row[tags_idx].lower() if tags_idx >= 0 and tags_idx < len(row) else ""
        tags_words = re.findall(r'[a-z][a-z0-9]*', tags_part)
        for w in tags_words:
            if len(w) < 3:
                continue
            if w.isdigit():
                continue
            if w in _STOPWORDS:
                continue
            if w in _PRODUCT_SUFFIXES:
                continue
            word_freq[w] += 1

    # 取最常见的词（排除型号后缀）
    sorted_words = [w for w, _ in word_freq.most_common(200)]
    product_keywords = [w for w in sorted_words if w not in _PRODUCT_SUFFIXES][:max_keywords]
    product_types = [w for w in sorted_words if w not in _STOPWORDS and w not in _PRODUCT_SUFFIXES][:20]

    # 生成自然语言描述
    top_words = [w for w, _ in word_freq.most_common(15)]
    all_text = " ".join(top_words)

    return {
        "keywords": product_keywords,
        "product_types": product_types,
        "top_handles": top_words,
        "all_text": all_text,
    }


def extract_hot_keywords_from_titles(csv_path, top_n=15, min_word_len=3):
    """从 CSV 的 Title 列统计热销关键词（按出现频率排序）

    返回: {
        "total_titles": int,
        "total_unique_words": int,
        "top_keywords": [(keyword, count), ...],  # 按频率降序
    }
    """
    import collections

    # 热销关键词停用词（只过滤纯营销/通用词/介词/单位词，不过滤产品属性词如 wireless/bluetooth）
    _HOT_KW_STOPWORDS = {
        "the", "a", "an", "and", "or", "for", "with", "by", "in", "on", "at",
        "to", "of", "from", "is", "are", "was", "were", "be", "been", "being",
        "has", "have", "had", "do", "does", "did", "will", "would", "could",
        "should", "may", "might", "must", "can", "this", "that", "these",
        "those", "it", "its", "all", "each", "every", "any", "some", "no",
        "not", "but", "if", "then", "so", "because", "as", "until", "while",
        "about", "after", "before", "above", "below", "between", "into",
        "through", "during", "out", "off", "over", "under", "again", "further",
        "once", "here", "there", "when", "where", "why", "how", "only", "same",
        "new", "sale", "buy", "shop", "store", "official", "online", "free",
        "shipping", "discount", "best", "top", "hot", "popular", "latest",
        "cheap", "price", "prices", "review", "reviews", "upc", "ean",
        "men", "women", "kids", "mens", "womens", "man", "woman", "boy", "girl",
        "without", "inch", "inches", "mm", "cm", "m", "ft", "watt", "watts", "v", "ah", "mah",
        "pack", "set", "piece", "pcs", "pair", "pairs", "box", "bag",
        "size", "sizes", "color", "colors", "colour", "colours", "type", "types",
        "model", "models", "style", "styles", "design", "designs",
        "quality", "high", "low", "good", "great", "nice", "perfect",
        "professional", "grade", "series", "generation", "gen",
        "usb", "cable", "charger", "battery", "power",
        "up", "down", "left", "right", "front", "back",
        "more", "less", "most", "least", "very", "really", "just",
        "than", "too", "also", "well", "much", "many", "few",
        "other", "another", "such", "like", "vs", "versus",
    }

    if not csv_path or not os.path.exists(csv_path):
        return None

    try:
        headers, rows = read_product_csv(csv_path)
    except Exception:
        return None

    if not headers or not rows:
        return None

    # 找到 Title 列索引
    title_idx = -1
    for i, h in enumerate(headers):
        if h.strip().lower() == "title":
            title_idx = i
            break

    if title_idx < 0:
        return None

    word_freq = collections.Counter()
    title_count = 0

    for row in rows:
        title = row[title_idx] if title_idx < len(row) else ""
        if not title:
            continue
        title_count += 1

        words = re.findall(r'[a-zA-Z]+', title.lower())
        seen = set()
        for w in words:
            if len(w) < min_word_len:
                continue
            if w.isdigit():
                continue
            if w in _HOT_KW_STOPWORDS:
                continue
            if w in seen:
                continue
            seen.add(w)
            word_freq[w] += 1

    top_keywords = word_freq.most_common(top_n)

    return {
        "total_titles": title_count,
        "total_unique_words": len(word_freq),
        "top_keywords": top_keywords,
    }


def generate_seo_title(domain, brand=None, product_keywords=None):
    """从 CSV 产品关键词生成 SEO title。

    完全基于 CSV 产品数据（product_keywords），不再依赖 title/description 输入。
    product_keywords 是 extract_product_keywords_from_csv() 的返回值。
    """
    if brand and brand.strip():
        brand_name = brand.strip()
    else:
        brand_name = "Our Store"

    domain_display = domain.replace("https://", "").replace("http://", "").split("/")[0]

    # 优先用 CSV 产品关键词决定品类词
    if product_keywords and product_keywords.get("product_types"):
        types = product_keywords["product_types"][:2]
        cat_kw = types[0] if types else _DEFAULTS["keyword"]
        cat_adj = _DEFAULTS["adjective"]
    else:
        cat_kw = _DEFAULTS["keyword"]
        cat_adj = _DEFAULTS["adjective"]

    candidates = []
    for tpl in _TITLE_TEMPLATES:
        for _ in range(2):
            try:
                filled = tpl.format(brand=brand_name, cat_kw=cat_kw.title(), cat_adj=cat_adj.title(), domain=domain_display)
                candidates.append(filled)
            except (KeyError, IndexError):
                continue

    valid = [c for c in candidates if len(c) <= 70
             and "discount" in c.lower()
             and not any(w.lower() in c.lower() for w in _AVOID_WORDS)]

    if not valid:
        base = f"{brand_name} Outlet | {cat_kw.title()} at Discount Prices"
        if len(base) > 70:
            base = f"{brand_name} Outlet | Discount {cat_kw.title()} Shop"
        valid.append(base[:70])

    valid.sort(key=lambda x: (len(x), random.random()), reverse=True)
    result = valid[0].strip()

    if len(result) > 70:
        result = result[:70]
    if "discount" not in result.lower():
        result = (result + " | Discount")[:70]

    return result.strip()


def generate_seo_description(domain, brand=None, product_keywords=None):
    """从 CSV 产品关键词生成 SEO description。

    完全基于 CSV 产品数据（product_keywords），不再依赖 title/description 输入。
    """
    if brand and brand.strip():
        brand_name = brand.strip()
    else:
        brand_name = "Our Store"

    # 优先用 CSV 产品关键词
    if product_keywords and product_keywords.get("product_types"):
        types = product_keywords["product_types"][:2]
        cat_kw = types[0] if types else _DEFAULTS["keyword"]
        benefit = _DEFAULTS["benefit"]
    else:
        cat_kw = _DEFAULTS["keyword"]
        benefit = _DEFAULTS["benefit"]

    candidates = []
    for tpl in _DESC_TEMPLATES:
        for _ in range(2):
            try:
                filled = tpl.format(brand=brand_name, cat_kw=cat_kw, benefit=benefit)
                candidates.append(filled)
            except (KeyError, IndexError):
                continue

    valid = [c for c in candidates if len(c) <= 320
             and not any(w.lower() in c.lower() for w in _AVOID_WORDS)]

    if not valid:
        valid.append(
            f"Shop {brand_name} Outlet Store for {cat_kw} at discount prices. "
            f"{benefit}. Free shipping worldwide."
        )

    valid.sort(key=lambda x: (len(x), random.random()), reverse=True)
    result = valid[0].strip()

    if len(result) > 320:
        result = result[:319]
    return result.strip()


# ===== Pages 内容生成 =====

def _load_pages(subdir="pages"):
    """扫描 data/{subdir}/ 目录下所有 .json 文件，按 handle 分组存储多版本
    返回: {handle: [page_def_1, page_def_2, ...]}
    同一个 handle 可能有多个 json 文件（如 about_us.json, about_us_v2.json），
    生成页面时会随机挑选一个版本，实现不重复效果"""
    script_dir = RESOURCE_PATH
    pages_dir = os.path.join(script_dir, "data", subdir)
    all_pages = {}
    if not os.path.isdir(pages_dir):
        return all_pages
    for fname in sorted(os.listdir(pages_dir)):
        if not fname.endswith('.json'):
            continue
        fpath = os.path.join(pages_dir, fname)
        try:
            with open(fpath, 'r', encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                pages = [data]
            elif isinstance(data, list):
                pages = data
            else:
                continue
            for page in pages:
                if 'handle' in page:
                    h = page['handle']
                    page['_source_file'] = fname
                    all_pages.setdefault(h, []).append(page)
        except Exception:
            continue
    return all_pages


def _safe_format(template, ctx):
    """安全地填充模板中的占位符。
    - 存在的 key → 正常替换
    - 不存在的 key → 用空字符串替代（防止 {xxx} 残留到最终页面）
    """
    if not template:
        return template
    class SafeDict(dict):
        def __missing__(self, key):
            return ""
    try:
        return template.format_map(SafeDict(ctx))
    except Exception:
        return template


def _render_page(page_def, ctx):
    """用品牌上下文渲染一个页面定义（支持 raw_html 或 sections 两种格式）"""
    try:
        if "raw_html" in page_def:
            raw = page_def["raw_html"]
            return _safe_format(raw, ctx)
        if "sections" in page_def:
            result_parts = []
            for section in page_def["sections"]:
                if "heading" in section:
                    result_parts.append(_safe_format(section["heading"], ctx))
                if "paragraphs" in section:
                    for p in section["paragraphs"]:
                        result_parts.append(_safe_format(p, ctx))
            return "".join(result_parts)
        return ""
    except Exception as e:
        return f"<!-- rendering error: {e} -->"


_PAGES_WITHOUT_DOMAIN = _load_pages("pages_without_domain")
_PAGES_WITH_DOMAIN = _load_pages("pages_with_domain")


def _get_pages(with_domain=False):
    """获取对应版本的页面模板集合"""
    return _PAGES_WITH_DOMAIN if with_domain else _PAGES_WITHOUT_DOMAIN


def _get_page_html(handle, ctx, page_def=None, with_domain=False):
    """渲染指定 handle 的页面 HTML
    - 如果传了 page_def，直接使用
    - 否则从对应版本的多个模板中随机选一个"""
    pages = _get_pages(with_domain)
    versions = pages.get(handle)
    if not versions:
        return ""
    if page_def is None:
        page_def = random.choice(versions)
    return _render_page(page_def, ctx)


def _get_page_title(handle, page_def=None, with_domain=False):
    """获取指定 handle 的页面标题
    - 如果传了 page_def，直接使用其 title
    - 否则从对应版本的多个模板中随机选一个"""
    pages = _get_pages(with_domain)
    versions = pages.get(handle)
    if not versions:
        return handle.replace("-", " ").title()
    if page_def is None:
        page_def = random.choice(versions)
    return page_def.get("title", handle.replace("-", " ").title())


def build_brand_context(seo_title, seo_description, domain, brand=None,
                         product_keywords=None, with_domain=False):
    """构建品牌上下文对象，供各页面生成函数使用。
    统一使用第一版占位符体系：domain / brand / email_*
    同时兼容 newtitle / new_description / newdomain 等旧版命名。
    product_keywords（dict）：来自 CSV 产品数据的核心关键词，注入到 ctx 供页面模板使用。
    with_domain: False=不带域名（{domain_link} 为纯文本），True=带域名（{domain_link} 为可点击链接）
    """
    domain_clean = domain.replace("https://", "").replace("http://", "").split("/")[0]
    if brand and brand.strip():
        brand_name = brand.strip()
    else:
        title_core = seo_title.split("|")[0].strip() if "|" in seo_title else seo_title
        title_core = title_core.replace("Outlet Store", "").replace("Outlet", "").replace("Discount Store", "").strip()
        brand_name = title_core
    category = detect_category(seo_title, seo_description)
    brand_value = brand_name if brand_name else "Our Store"
    brand_store_value = brand_name + " Outlet Store" if brand_name and "Outlet" not in brand_name else seo_title.split("|")[0].strip()

    # 根据 with_domain 决定 domain_link 的格式
    if with_domain:
        domain_link_html = f'<a href="https://{domain_clean}" target="_blank" style="color: inherit; text-decoration: underline;">{domain_clean}</a>'
    else:
        domain_link_html = domain_clean  # 纯文本，无链接

    ctx = {
        # 第一版标准占位符
        "domain": domain_clean,
        "title": seo_title,
        "description": seo_description,
        "brand": brand_value,
        "brand_store": brand_store_value,
        "domain_link": domain_link_html,  # 带域名=HTML链接，不带域名=纯文本
        "email_press": f"press@{domain_clean}",
        "email_support": f"support@{domain_clean}",
        "email_wholesale": f"wholesale@{domain_clean}",
        "category": category,
        # CSV 产品关键词（用于 SEO title/description 生成，不用于页面内容）
        "product_keywords": product_keywords or {},
        # 兼容性：旧版命名的备用 key
        "newtitle": brand_value,
        "new_description": seo_description,
        "newdomain": domain_clean,
        "brand_name": brand_value,
    }
    return ctx


def generate_all_content(domain, brand=None, product_csv_path=None, with_domain=False):
    """主入口：生成新 SEO title/description + 完整页面内容。

    完全基于 CSV 产品数据，不再依赖 title/description 输入。
    product_csv_path：CSV 产品数据路径，用于提取产品关键词并生成 SEO 内容和页面内容。
    with_domain: False=不带域名版本（纯文本），True=带域名版本（可点击链接）
    """
    # 从 CSV 提取产品关键词
    product_keywords = extract_product_keywords_from_csv(product_csv_path) if product_csv_path else None

    new_seo_title = generate_seo_title(
        domain, brand=brand, product_keywords=product_keywords
    )
    new_seo_description = generate_seo_description(
        domain, brand=brand, product_keywords=product_keywords
    )
    ctx = build_brand_context(
        new_seo_title, new_seo_description, domain, brand=brand,
        product_keywords=product_keywords, with_domain=with_domain
    )

    pages = {}
    for handle, versions in _get_pages(with_domain).items():
        # 从该 handle 的多个版本中随机选一个，保证 title 和 body_html 同源
        page_def = random.choice(versions)
        pages[handle] = {
            "handle": handle,
            "title": _get_page_title(handle, page_def=page_def, with_domain=with_domain),
            "body_html": _get_page_html(handle, ctx, page_def=page_def, with_domain=with_domain),
        }

    menus = [
        {
            "menu_handle": "main-menu",
            "menu_title": "Main Menu",
            "is_default": "TRUE",
            "items": [],
        },
    ]

    return {
        "seo_title": new_seo_title,
        "seo_description": new_seo_description,
        "pages": pages,
        "menus": menus,
        "context": ctx,
    }


# ============================================================
# 固定列标题（Altera 导入标准，完全不依赖任何模板文件）
# ============================================================

METAFIELD_DEFINITION_COLUMNS = [
    "Namespace", "Key", "Command", "Name", "Owner Type", "Description",
    "Pinned: Is Pinned", "Pinned: Position",
    "Type: Category", "Type: Name",
    "Validation: Name", "Validation: Type", "Validation: Value",
]

SMART_COLLECTION_COLUMNS = [
    "Handle", "Command", "Title", "Body HTML", "Sort Order",
    "Published", "Published Scope", "Row #", "Top Row",
    "Image Src", "Image Width", "Image Height", "Image Alt Text",
    "Must Match", "Rule: Product Column", "Rule: Relation", "Rule: Condition",
    "Published: Online Store", "Published: POS", "Published: Shop",
]

MENU_COLUMNS = [
    "ID", "Handle", "Command", "Title", "Is Default", "Top Row", "Row #",
    "Menu Item: ID", "Menu Item: Title", "Menu Item: Command",
    "Menu Item: Resource Type", "Menu Item: Resource ID",
    "Menu Item: Resource Handle", "Menu Item: Collection Tags",
    "Menu Item: URL", "Menu Item: Parent ID", "Menu Item: Parent Title",
    "Menu Item: Position",
]

PAGES_COLUMNS = ["Handle", "Title", "Body (HTML)", "Template Suffix", "Published"]


# ============================================================
# 工具函数：CSV 解析 → Metafield Definitions
# ============================================================

def parse_column_name(col):
    """解析列名 → (name, cf_suffix)

    示例:
      'Description (product.metafields.c_f.Description)'
        → ('Description', 'c_f.Description')
      'Reviews(product.metafields.c_f.Reviews)'
        → ('Reviews', 'c_f.Reviews')
    """
    match = re.match(r'(.+?)\s*\((.+)\)', col)
    if match:
        name = match.group(1).strip()
        content = match.group(2)
        # 匹配 c_f.xxx
        cf_match = re.search(r'c_f\.\w+', content)
        if cf_match:
            return name, cf_match.group(0)
        # 匹配 product.metafields.<namespace>.<key>
        mf_match = re.search(r'product\.metafields\.(\w+)\.(\w+)', content)
        if mf_match:
            return name, f"{mf_match.group(1)}.{mf_match.group(2)}"
    return col.strip(), None


def get_metafield_columns(input_csv):
    """从 CSV 中找出所有 metafield 列并解析。

    返回: list[(column_name, cf_suffix)]
    """
    with open(input_csv, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return []
        columns = reader.fieldnames

    metafield_cols = []
    for col in columns:
        if "c_f" in col or "product.metafields." in col:
            name, cf_suffix = parse_column_name(col)
            if cf_suffix:
                metafield_cols.append((name, cf_suffix))
    return metafield_cols


def build_metafield_definitions_from_csv(input_csv):
    """根据产品 CSV 生成 Metafield Definitions 行。

    逻辑:
      1. 扫描列标题，找出含 'c_f' 或 'product.metafields.' 的列
      2. 用 parse_column_name 解析出 name 和 namespace.key
      3. 每列生成一条 Metafield Definition 记录
    """
    metafield_cols = get_metafield_columns(input_csv)
    rows = []
    for row_number, (name, cf_suffix) in enumerate(metafield_cols, start=1):
        namespace, key = cf_suffix.split(".", 1)
        rows.append({
            "Namespace": namespace,
            "Key": key,
            "Command": "MERGE",
            "Name": name,
            "Owner Type": "PRODUCT",
            "Description": "",
            "Pinned: Is Pinned": "TRUE",
            "Pinned: Position": str(row_number),
            "Type: Category": "TEXT",
            "Type: Name": "multi_line_text_field",
            "Validation: Name": "",
            "Validation: Type": "",
            "Validation: Value": "",
        })
    return rows


# ============================================================
# 工具函数：生成 Shopify custom block ID（custom_xxxxxx 格式）
# ============================================================

def generate_custom_block_id(existing_ids=None):
    """生成 Shopify 风格的 custom block ID: custom_ + 6位随机字母数字"""
    chars = string.ascii_letters + string.digits
    existing = existing_ids or set()
    for _ in range(100):
        suffix = "".join(random.choices(chars, k=6))
        bid = f"custom_{suffix}"
        if bid not in existing:
            return bid
    suffix = "".join(random.choices(chars, k=10))
    return f"custom_{suffix}"


# ============================================================
# 工具函数：修改 product.json，根据 metafield 列表添加自定义 tabs
# ============================================================

def _find_product_main_section(data):
    """在 product.json 的 sections 中找到产品主 section（包含 description tab 的那个）。

    遍历所有 section，找到第一个包含 type=description block 的 section。
    找不到的话返回 None。
    """
    for sec_id, section in data.get("sections", {}).items():
        blocks = section.get("blocks", {})
        if not isinstance(blocks, dict):
            continue
        for bid, block in blocks.items():
            if isinstance(block, dict) and block.get("type") == "description":
                return sec_id, section
    return None, None


def build_custom_tabs_in_product_json(product_json_path, metafield_rows):
    """修改 product.json，根据 Metafield Definitions 添加自定义 tabs。

    - 移除原有的 type=custom 的 blocks
    - 为每个 metafield 生成一个 custom tab block
    - title = Name.upper()（全部大写）
    - product_tab_key_metafield = Key
    - 插入到 description tab 之后，share tab 之前
    - 通过 type=description block 定位主 section，不依赖固定 section id
    """
    with open(product_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    sec_id, main = _find_product_main_section(data)
    if main is None:
        return

    blocks = main["blocks"]
    block_order = main["block_order"]

    # 1. 移除原有的 type=custom 的 blocks
    custom_ids = [bid for bid, b in blocks.items() if b.get("type") == "custom"]
    for bid in custom_ids:
        del blocks[bid]
    block_order = [bid for bid in block_order if bid not in custom_ids]

    # 2. 找到插入位置：description tab 之后，share tab 之前
    insert_idx = len(block_order)  # 默认插末尾
    for i, bid in enumerate(block_order):
        if blocks.get(bid, {}).get("type") == "description":
            insert_idx = i + 1
            break

    # 3. 生成新的 custom tabs
    existing_ids = set(blocks.keys())
    for mf in metafield_rows:
        name = mf.get("Name", "")
        key = mf.get("Key", "")
        bid = generate_custom_block_id(existing_ids)
        existing_ids.add(bid)

        blocks[bid] = {
            "type": "custom",
            "settings": {
                "open_tab_mobile": False,
                "title": name.upper(),
                "type": "metafield",
                "content": "",
                "product_custom_source": "key_text",
                "product_tab_key_metafield": key,
                "enable_btn_show_more": False,
                "maximum_des_to_show": 300,
                "zdy_biantihezi": False,
            },
        }
        block_order.insert(insert_idx, bid)
        insert_idx += 1

    main["block_order"] = block_order

    with open(product_json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))


# ============================================================
# 工具函数：强制删除目录（带重试，解决 Windows 权限/占用问题）
# ============================================================

def _force_rmtree(path):
    """强制删除目录，带重试机制，处理 Windows 下文件占用和权限问题。"""
    import time
    def _onerror(func, p, exc_info):
        try:
            os.chmod(p, 0o777)
        except Exception:
            pass
        try:
            func(p)
        except Exception:
            pass
    for attempt in range(3):
        try:
            shutil.rmtree(path, onerror=_onerror)
            return
        except Exception:
            if attempt < 2:
                time.sleep(0.5)
    try:
        shutil.rmtree(path, ignore_errors=True)
    except Exception:
        pass


# ============================================================
# 工具函数：复制主题模板 → 修改 product.json → 打包 zip
# ============================================================

def build_theme_zip(brand, domain, metafield_rows, output_dir, hot_keywords=None):
    """根据品牌生成定制化的主题 zip 包。

    1. 复制 data/shopify模板 到输出目录
    2. 修改 templates/product.json，添加自定义 tabs + 随机销量数
    3. 修改 config/settings_data.json，随机销量/浏览数 + 热搜词
    4. 修改 sections/footer-group.json，替换版权域名
    5. 打包成 zip
    6. 返回 zip 文件路径
    """
    src_theme = os.path.join(RESOURCE_PATH, "data", "shopify模板")
    safe_brand = "".join(c for c in brand if c.isalnum() or c in "-_")[:40] or "store"
    theme_folder = os.path.join(output_dir, f"{safe_brand}_theme")
    zip_path = os.path.join(output_dir, f"{safe_brand}_theme.zip")

    # 1. 复制主题文件夹
    if os.path.exists(theme_folder):
        _force_rmtree(theme_folder)
    shutil.copytree(src_theme, theme_folder)

    # 2. 修改 product.json
    product_json = os.path.join(theme_folder, "templates", "product.json")
    if os.path.exists(product_json):
        build_custom_tabs_in_product_json(product_json, metafield_rows)
        randomize_sold_in_number(product_json)

    # 3. 修改 settings_data.json
    settings_json = os.path.join(theme_folder, "config", "settings_data.json")
    if os.path.exists(settings_json):
        randomize_settings_data(settings_json, hot_keywords=hot_keywords)

    # 4. 修改 footer-group.json
    footer_json = os.path.join(theme_folder, "sections", "footer-group.json")
    if os.path.exists(footer_json):
        set_footer_link_list_menus(footer_json)
        if domain:
            replace_footer_copyright_domain(footer_json, domain)

    # 5. 打包 zip
    if os.path.exists(zip_path):
        os.remove(zip_path)

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(theme_folder):
            for fn in files:
                full = os.path.join(root, fn)
                arcname = os.path.relpath(full, theme_folder)
                zf.write(full, arcname)

    # 6. 清理临时文件夹
    _force_rmtree(theme_folder)

    return zip_path


# ============================================================
# 工具函数：生成随机数字符串（逗号分隔）
# ============================================================

def _random_numbers(count, min_val, max_val, with_space=False):
    """生成 count 个 [min_val, max_val] 范围内的随机数，逗号分隔。

    with_space=True 时，逗号后加空格（与 customer_viewing_number 格式一致）。
    """
    nums = [str(random.randint(min_val, max_val)) for _ in range(count)]
    sep = ", " if with_space else ","
    return sep.join(nums)


# ============================================================
# 工具函数：product.json - 随机销量数字
# ============================================================

def randomize_sold_in_number(product_json_path):
    """给 product.json 中所有 type=meta 的 block 的 sold_in_number 生成 10 个左右 200~500 的随机数。

    用 type 定位，不依赖固定 section id 和 block id。
    遍历所有 section 的所有 block，找到 type=meta 的进行修改。
    """
    with open(product_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    for sec_id, section in data.get("sections", {}).items():
        blocks = section.get("blocks", {})
        if not isinstance(blocks, dict):
            continue
        for bid, block in blocks.items():
            if isinstance(block, dict) and block.get("type") == "meta":
                settings = block.get("settings", {})
                if "sold_in_number" in settings:
                    settings["sold_in_number"] = _random_numbers(10, 200, 500)
                    block["settings"] = settings

    with open(product_json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))


# ============================================================
# 工具函数：settings_data.json - 随机销量 + 浏览数 + 热搜词
# ============================================================

def randomize_settings_data(settings_json_path, hot_keywords=None):
    """给 settings_data.json 生成随机的 quick_view 销量、浏览数和热搜词。

    - quick_view_sold_in_number: 10 个 200~500 随机数
    - quick_view_customer_viewing_number: 20 个 100~350 随机数
    - seach_trending_item_1~6: 前6个热销关键词（从产品标题统计）
    """
    sold_val = _random_numbers(10, 200, 500)
    view_val = _random_numbers(20, 100, 350, with_space=True)

    with open(settings_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    current = data.get("current", {})

    if "quick_view_sold_in_number" in current:
        current["quick_view_sold_in_number"] = sold_val

    if "quick_view_customer_viewing_number" in current:
        current["quick_view_customer_viewing_number"] = view_val

    if hot_keywords and hot_keywords.get("top_keywords"):
        kw_list = [kw for kw, cnt in hot_keywords["top_keywords"]]
        for i in range(1, 7):
            key = f"seach_trending_item_{i}"
            if key in current and i <= len(kw_list):
                current[key] = kw_list[i - 1]

    data["current"] = current

    with open(settings_json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

    schema_path = os.path.join(os.path.dirname(settings_json_path), "settings_schema.json")
    if os.path.exists(schema_path):
        _update_schema_defaults(schema_path, {
            "quick_view_sold_in_number": sold_val,
            "quick_view_customer_viewing_number": view_val,
        })


def _update_schema_defaults(schema_path, values_map):
    """更新 settings_schema.json 中指定 id 的 default 值。"""
    with open(schema_path, "r", encoding="utf-8") as f:
        schema = json.load(f)

    def _walk(obj):
        if isinstance(obj, dict):
            if "id" in obj and "default" in obj and obj["id"] in values_map:
                obj["default"] = values_map[obj["id"]]
            for v in obj.values():
                if isinstance(v, (dict, list)):
                    _walk(v)
        elif isinstance(obj, list):
            for item in obj:
                if isinstance(item, (dict, list)):
                    _walk(item)

    _walk(schema)

    with open(schema_path, "w", encoding="utf-8") as f:
        json.dump(schema, f, ensure_ascii=False, indent=2)


# ============================================================
# 工具函数：footer-group.json - 替换版权里的域名
# ============================================================

def replace_footer_copyright_domain(footer_json_path, domain):
    """遍历 footer-group.json，把所有 copyright_content 里的域名替换成输入的域名。

    不依赖固定路径，递归查找所有 copyright_content 字段，用正则替换里面的域名部分。
    支持多种格式：Copyright 2026©domain、Copyright 2026 © domain、© Copyright 2026 domain 等。
    """
    with open(footer_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    domain_re = re.compile(
        r"(?:[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?\.)+"
        r"(?:com|net|org|cn|io|co|shop|store|info|biz|xyz|top|vip|me|tv|cc)"
        r"(?:\.[a-zA-Z]{2})?",
        re.IGNORECASE
    )

    def _replace_in_obj(obj):
        if isinstance(obj, dict):
            for k, v in obj.items():
                if k == "copyright_content" and isinstance(v, str):
                    if domain_re.search(v):
                        obj[k] = domain_re.sub(domain, v)
                    else:
                        obj[k] = re.sub(
                            r"(Copyright\s*\d{4}\s*©\s*)[^<\s]+",
                            lambda m: m.group(1) + domain,
                            v
                        )
                else:
                    _replace_in_obj(v)
        elif isinstance(obj, list):
            for item in obj:
                _replace_in_obj(item)

    _replace_in_obj(data)

    with open(footer_json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))


# ============================================================
# 工具函数：footer-group.json - 设置 link_list 的 menu
# ============================================================

FOOTER_LINK_LIST_MENUS = ["main-menu", "policy", "about"]

def set_footer_link_list_menus(footer_json_path):
    """给 footer-group.json 中所有 type=link_list 的 block 按顺序设置 menu。

    按 block_order 中出现的顺序，依次设置为 main-menu、shop、policy。
    用 type 定位，不依赖固定 block id。
    """
    with open(footer_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    for section_id, section in data.get("sections", {}).items():
        blocks = section.get("blocks", {})
        block_order = section.get("block_order", [])
        if not blocks or not block_order:
            continue

        link_list_idx = 0
        for bid in block_order:
            block = blocks.get(bid, {})
            if block.get("type") == "link_list":
                if link_list_idx < len(FOOTER_LINK_LIST_MENUS):
                    settings = block.get("settings", {})
                    settings["menu"] = FOOTER_LINK_LIST_MENUS[link_list_idx]
                    block["settings"] = settings
                    link_list_idx += 1

    with open(footer_json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))


# ============================================================
# 工具函数：大文件 CSV 分割（按产品 ID 完整性，不超过 14MB）
# ============================================================

def split_product_csv(input_csv, brand, output_dir, max_mb=12):
    """按产品 ID 完整性切分 CSV，确保同一 ID 的所有行在同一文件中，
    每个文件严格不超过 max_mb（默认 12MB，预留余量确保 <14MB）。

    返回生成的文件路径列表。
    """
    safe_brand = "".join(c for c in brand if c.isalnum() or c in "-_")[:40] or "store"
    file_size = os.path.getsize(input_csv)

    # 第一步：读取所有行并按第一列（ID）分组
    rows_by_id = {}
    id_order = []
    header = None
    total_rows = 0

    with open(input_csv, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            total_rows += 1
            pid = row[0] if row else ""
            if pid not in rows_by_id:
                rows_by_id[pid] = []
                id_order.append(pid)
            rows_by_id[pid].append(row)

    # 计算平均每行字节数（估算用）
    avg_row_bytes = file_size / max(total_rows, 1)
    target_rows = int((max_mb * 1024 * 1024) / avg_row_bytes)

    # 第二步：按 ID 分组打包
    output_files = []
    current_ids = []
    current_rows = 0
    part_num = 1

    for pid in id_order:
        pid_row_count = len(rows_by_id[pid])

        if current_rows + pid_row_count > target_rows and current_ids:
            # 保存当前批次
            out_path = _save_csv_chunk(
                header, rows_by_id, current_ids,
                output_dir, safe_brand, part_num
            )
            output_files.append(out_path)
            part_num += 1
            current_ids = []
            current_rows = 0

        current_ids.append(pid)
        current_rows += pid_row_count

    if current_ids:
        out_path = _save_csv_chunk(
            header, rows_by_id, current_ids,
            output_dir, safe_brand, part_num
        )
        output_files.append(out_path)

    return output_files


def _save_csv_chunk(header, rows_by_id, ids, output_dir, brand, part_num):
    """保存一个 CSV 分片，返回文件路径。"""
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{brand}_split_part_{part_num}.csv"
    out_path = os.path.join(output_dir, filename)

    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for pid in ids:
            for row in rows_by_id[pid]:
                writer.writerow(row)

    return out_path


# ============================================================
# 工具函数：CSV Tags 解析 → Smart Collections
# ============================================================

def make_handle(title):
    """标题 → Shopify handle（小写、短横线连接）"""
    handle = title.strip().lower()
    handle = re.sub(r"[^a-z0-9]+", "-", handle)
    return handle.strip("-")


def get_unique_tags_from_csv(input_csv):
    """从 CSV 的 Tags 列提取所有去重后的 tag（保持出现顺序）"""
    unique_tags = []
    seen = set()
    with open(input_csv, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames or "Tags" not in reader.fieldnames:
            return []
        for row in reader:
            tags_str = row.get("Tags", "").strip()
            if not tags_str:
                continue
            for tag in tags_str.split(","):
                tag = tag.strip()
                if not tag:
                    continue
                key = tag.casefold()
                if key in seen:
                    continue
                seen.add(key)
                unique_tags.append(tag)
    return unique_tags


def build_smart_collections_from_csv(input_csv):
    """从产品 CSV 的 Tags 列生成 Smart Collections 行。

    严格按参考逻辑：
      - 每个 unique tag = 1 条 collection（1 行）
      - **每行 Top Row = TRUE**（每个 collection 都是独立记录，
        Altera 不会把后续行当成前一个 collection 的附加条件）
      - Rule: Product Column = Tag, Relation = Equals, Condition = 原始 tag
      - 不做任何前缀补全（由 generate_prefix_collections 处理）
    """
    tags = get_unique_tags_from_csv(input_csv)
    if not tags:
        return []

    rows = []
    for row_number, tag in enumerate(tags, start=1):
        clean_tag = normalize_text(tag)
        rows.append({
            "Handle": make_handle(tag),
            "Command": "MERGE",
            "Title": clean_tag,
            "Body HTML": "",
            "Sort Order": "Best Selling",
            "Published": "TRUE",
            "Published Scope": "global",
            "Row #": row_number,  # int 而非 str（Excel 中以数字显示）
            "Top Row": "TRUE",  # 关键：每个 collection 行都是新记录的开始
            "Image Src": "",
            "Image Width": "",
            "Image Height": "",
            "Image Alt Text": "",
            "Must Match": "all conditions",
            "Rule: Product Column": "Tag",
            "Rule: Relation": "Equals",
            "Rule: Condition": clean_tag,
            "Published: Online Store": "TRUE",
            "Published: POS": "TRUE",
            "Published: Shop": "TRUE",
        })
    return rows


def _make_smart_collection_row(title, relation="Equals", condition="", row_number=0):
    """创建一条 Smart Collection 行。

    注意：Shopify 对 Tag 字段仅支持 Equals / Not Equals，不支持 Contains。
    对于 prefix collections（菜单分组节点），传空的 relation 与 condition
    会生成一个手动集合（manual collection），可安全导入。
    """
    handle = make_handle(normalize_text(title))
    return {
        "Handle": handle,
        "Command": "MERGE",
        "Title": normalize_text(title),
        "Body HTML": "",
        "Sort Order": "Best Selling",
        "Published": "TRUE",
        "Published Scope": "global",
        "Row #": str(row_number),
        "Top Row": "",
        "Image Src": "",
        "Image Width": "",
        "Image Height": "",
        "Image Alt Text": "",
        "Must Match": "all conditions" if condition else "",
        "Rule: Product Column": "Tag" if condition else "",
        "Rule: Relation": relation if condition else "",
        "Rule: Condition": condition,
        "Published: Online Store": "TRUE",
        "Published: POS": "TRUE",
        "Published: Shop": "TRUE",
    }



# ============================================================
# 工具函数：Smart Collections → Menu（含父子层级）
# ============================================================

# ============================================================
# 菜单工具函数（按参考代码逻辑重写）
# ============================================================

_CONJUNCTIONS = {
    # 并列连词
    "&", "and", "or", "+",
    # 介词（基础）
    "with", "to", "by", "from", "in", "for", "of",
    "on", "at", "off", "over", "under", "out", "up", "as", "per",
    # 介词（扩展：价格范围/时间类）
    "after", "before", "above", "below", "without",
    # 对比/附加
    "vs", "versus", "plus",
}

# ============================================================
# 产品型号后缀词 — 当以下词出现在标题末尾时，
# 表示这是同一产品线的不同型号，不是分类层级。
# 例如 "Ekster Cardholder Pro" 不是 "Ekster Cardholder → Pro" 的父子关系，
# 而是 4 个并列的独立产品型号。
# ============================================================
_MODEL_SUFFIXES = {
    # 常见尺寸/型号
    "pro", "max", "mini", "air", "slim", "xl", "xxl", "lg", "sm",
    # 等级/版本
    "plus", "standard", "premium", "lite", "classic", "basic",
    "advanced", "ultra", "essential", "ultimate", "elite", "core",
    # 风格/年代
    "sport", "fit", "active", "original", "edition", "v2", "v3",
    # 材料
    "leather", "nylon", "canvas", "aluminum", "steel",
    # 功能
    "smart", "digital", "wireless", "bluetooth",
    # 数量/大小
    "small", "large", "medium",
    # 常用后缀（带数字的型号，如 iPhone 14, Series 7）
}

def _is_model_suffix(word):
    """判断一个词是否是产品型号后缀。
    规则：
      - 在 _MODEL_SUFFIXES 中 → 是
      - 1-4 位纯数字（如 14, 7, 2024）→ 是
      - 5 位以上纯数字（如 1774, 10086）→ 不是（更像是分类编号）
      - 字母+数字混合（如 S24, XR15, A15）→ 是
    """
    if not word:
        return False
    w = word.lower()
    if w in _MODEL_SUFFIXES:
        return True
    if w.isdigit():
        return len(w) <= 3
    # 字母+数字混合：至少 1 个字母 + 至少 1 个数字
    if any(c.isalpha() for c in w) and any(c.isdigit() for c in w):
        return True
    return False


_MAX_MENU_DEPTH = 3
_MIN_INFERRED_GROUP_WORDS = 2
_MIN_INFERRED_GROUP_CHILDREN = 2


def normalize_text(value):
    """通用文本规范化：去首尾空白、压缩多余空白"""
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def normalize_title_for_lookup(title):
    """标题规范化：给 & + 等符号周围加空格，方便按单词拆分"""
    return normalize_text(str(title).replace("&", " & ").replace("+", " + "))


def title_words(title):
    return normalize_title_for_lookup(title).split()


def words_to_title(words):
    return normalize_text(" ".join(words))


def build_title_map(titles, ids):
    """构建 norm_title → menu_id 映射，同时返回重复标题集合"""
    title_map = {}
    duplicates = set()
    for title, menu_id in zip(titles, ids):
        norm = normalize_title_for_lookup(title)
        if not norm:
            continue
        if norm in title_map:
            duplicates.add(title)
            continue
        title_map[norm] = menu_id
    return title_map, duplicates


def find_smart_parent(current_title, title_map):
    """用最长可用前缀作为父级。
    返回父级 menu_id（0 表示没有父级，顶级）"""
    parts = title_words(current_title)
    if len(parts) <= 1:
        return 0

    for i in range(len(parts) - 1, 0, -1):
        potential_parent = words_to_title(parts[:i])
        if potential_parent not in title_map:
            continue

        return title_map[potential_parent]

    return 0


def is_valid_parent_boundary(words, prefix_length):
    """避免父级标题以连词结尾"""
    if prefix_length <= 0 or prefix_length >= len(words):
        return False
    last_word = words[prefix_length - 1].lower()
    return last_word not in _CONJUNCTIONS


def find_existing_parent_word_count(words, existing_titles):
    """对一个候选前缀，看它自己是否有更上层的父级已存在（返回父级单词数）"""
    for i in range(len(words) - 1, 0, -1):
        potential_parent = words_to_title(words[:i])
        if potential_parent in existing_titles:
            return i
    return 0


def infer_missing_group_titles(titles):
    """从已有标题推断缺失的中间层级。

    严格约束：
      1. 标题单词数 > 2 才参与推断
      2. 候选前缀不能已经是 collection（已存在的标题）
      3. 候选前缀必须有一个更上层的父级已存在
      4. 前缀与父级的单词差 >= MIN_INFERRED_GROUP_WORDS
      5. 该前缀作为父级统辖的子项 >= MIN_INFERRED_GROUP_CHILDREN 才采用
    """
    from collections import defaultdict

    clean_titles = [normalize_text(t) for t in titles if normalize_text(t)]
    existing_titles = {
        normalize_title_for_lookup(t)
        for t in clean_titles
        if normalize_title_for_lookup(t)
    }

    prefix_children = defaultdict(set)
    title_word_lists = []

    for title in clean_titles:
        words = title_words(title)
        if len(words) <= 2:
            continue
        title_key = words_to_title(words)
        title_word_lists.append(words)

        for prefix_length in range(1, len(words)):
            if not is_valid_parent_boundary(words, prefix_length):
                continue

            prefix_title = words_to_title(words[:prefix_length])
            if prefix_title in existing_titles:
                continue

            parent_word_count = find_existing_parent_word_count(
                words[:prefix_length],
                existing_titles,
            )
            if parent_word_count == 0:
                continue
            if prefix_length - parent_word_count < _MIN_INFERRED_GROUP_WORDS:
                continue

            prefix_children[tuple(words[:prefix_length])].add(title_key)

    candidate_prefixes = {
        prefix_words: len(children)
        for prefix_words, children in prefix_children.items()
        if len(children) >= _MIN_INFERRED_GROUP_CHILDREN
    }

    inferred_groups = set()
    for words in title_word_lists:
        matching_prefixes = [
            prefix_words
            for prefix_words in candidate_prefixes
            if len(prefix_words) < len(words)
            and tuple(words[:len(prefix_words)]) == prefix_words
        ]
        if matching_prefixes:
            best_prefix = max(
                matching_prefixes,
                key=lambda prefix_words: (
                    candidate_prefixes[prefix_words],
                    len(prefix_words),
                ),
            )
            inferred_groups.add(words_to_title(best_prefix))

    return sorted(inferred_groups, key=str.casefold)


def trim_parent_prefix(current_title, parent_title):
    """子项标题裁剪父级前缀"""
    current = normalize_text(current_title)
    parent = normalize_text(parent_title)
    if not parent:
        return current
    prefix = f"{parent} "
    if current.startswith(prefix):
        return current[len(prefix):].strip()
    return current


def build_title_handle_map(collection_rows):
    """构建 norm_title → handle 映射"""
    title_handle_map = {}
    for col in collection_rows:
        title = normalize_text(col.get("Title", ""))
        handle = normalize_text(col.get("Handle", ""))
        if not title:
            continue
        key = normalize_title_for_lookup(title)
        if key and key not in title_handle_map:
            title_handle_map[key] = handle
    return title_handle_map


def find_parent_handle_for_inferred_title(inferred_title, title_handle_map):
    """对推断的中间节点，找最长存在的前缀的 handle"""
    words = title_words(inferred_title)
    for i in range(len(words) - 1, 0, -1):
        parent_title = words_to_title(words[:i])
        handle = title_handle_map.get(parent_title, "")
        if handle:
            return handle
    return ""


def build_menu_from_collections(collection_rows):
    """根据 Smart Collections 构建菜单（与生成菜单完整.py 逻辑完全一致）。"""
    if not collection_rows:
        return []

    MENU_OWN_ID = "100000000001"

    # ---- 1. 提取并去重基础条目（按 Title + Handle 组合去重）----
    base_rows = []
    seen_pairs = set()
    for col in collection_rows:
        title = normalize_text(col.get("Title", ""))
        if not title:
            continue
        handle = normalize_text(col.get("Handle", ""))
        if not handle:
            handle = make_handle(title)
        pair_key = (title, handle)
        if pair_key in seen_pairs:
            continue
        seen_pairs.add(pair_key)
        base_rows.append({"title": title, "handle": handle})

    if not base_rows:
        return []

    # ---- 2. 推断缺失的中间层级 ----
    all_titles = [r["title"] for r in base_rows]
    inferred_titles = infer_missing_group_titles(all_titles)

    # ---- 3. 为推断的标题追加行（Handle 复用最近父级的 handle）----
    title_handle_map = build_title_handle_map(collection_rows)
    seen_titles = set(r["title"] for r in base_rows)
    for title in inferred_titles:
        norm_title = normalize_text(title)
        if not norm_title:
            continue
        if norm_title in seen_titles:
            continue
        handle = find_parent_handle_for_inferred_title(norm_title, title_handle_map)
        if not handle:
            continue
        seen_titles.add(norm_title)
        base_rows.append({"title": norm_title, "handle": handle})

    # ---- 4. 按 Title 字母升序排序 ----
    base_rows.sort(key=lambda r: r["title"].casefold())
    n = len(base_rows)
    if n == 0:
        return []
    seq_ids = list(range(1, n + 1))

    full_titles = [r["title"] for r in base_rows]
    handles = [r["handle"] for r in base_rows]

    # ---- 5. 计算父子关系 ----
    title_map, _ = build_title_map(full_titles, seq_ids)
    parent_ids = []
    for idx, title in enumerate(full_titles):
        my_id = seq_ids[idx]
        parent_id = find_smart_parent(title, title_map)
        if parent_id == my_id:
            parent_id = 0
        parent_ids.append(parent_id)

    # ---- 6. clamp_depth：超过 MAX_DEPTH 的节点逐级上提 ----
    parent_map = dict(zip(seq_ids, parent_ids))

    def _get_level_with_memo(menu_id, memo, visiting):
        if menu_id in memo:
            return memo[menu_id]
        if menu_id in visiting:
            memo[menu_id] = 1
            return 1
        visiting.add(menu_id)
        pid = parent_map.get(menu_id, 0)
        if pid == 0 or pid not in parent_map:
            level = 1
        else:
            level = 1 + _get_level_with_memo(pid, memo, visiting)
        visiting.discard(menu_id)
        memo[menu_id] = level
        return level

    def _calc_all_levels():
        memo = {}
        for mid in parent_map:
            _get_level_with_memo(mid, memo, set())
        return memo

    # 迭代上提，最多迭代 100 次
    for _ in range(100):
        changed = False
        level_map = _calc_all_levels()
        for mid in list(parent_map.keys()):
            if level_map.get(mid, 1) <= _MAX_MENU_DEPTH:
                continue
            p = parent_map.get(mid, 0)
            grandparent = parent_map.get(p, 0) if p else 0
            if grandparent == mid:
                parent_map[mid] = 0
            else:
                parent_map[mid] = grandparent
            changed = True
        if not changed:
            break

    parent_ids = [parent_map.get(mid, 0) for mid in seq_ids]

    # ---- 7. trim_parent_prefix 裁剪子项显示标题 ----
    full_title_map = dict(zip(seq_ids, full_titles))
    display_titles = []
    for title, pid in zip(full_titles, parent_ids):
        parent_title = full_title_map.get(pid, "") if pid else ""
        display_titles.append(trim_parent_prefix(title, parent_title))
    display_title_map = dict(zip(seq_ids, display_titles))

    # ---- 8. 计算 Parent Title ----
    parent_titles = []
    for pid in parent_ids:
        if pid and pid in display_title_map:
            parent_titles.append(display_title_map[pid])
        else:
            parent_titles.append("")

    # ---- 9. 计算 Position（同父级下从 1 开始连续编号）----
    pos_counter = {}
    position_list = []
    for pid in parent_ids:
        key = str(pid) if pid else "0"
        pos_counter[key] = pos_counter.get(key, 0) + 1
        position_list.append(str(pos_counter[key]))

    # ---- 10. 生成 menu_rows ----
    menu_rows = []
    for idx in range(n):
        mid = seq_ids[idx]
        pid = parent_ids[idx]
        display_title = display_titles[idx]
        parent_title = parent_titles[idx]
        handle = handles[idx] if handles[idx] else make_handle(full_titles[idx])

        parent_id_str = str(pid) if pid else ""

        menu_rows.append({
            "ID": MENU_OWN_ID,
            "Handle": "main-menu",
            "Command": "MERGE",
            "Title": "Main menu",
            "Is Default": "TRUE",
            "Top Row": "TRUE" if idx == 0 else "",
            "Row #": str(idx + 1),
            "Menu Item: ID": str(mid),
            "Menu Item: Title": display_title,
            "Menu Item: Command": "MERGE",
            "Menu Item: Resource Type": "COLLECTION",
            "Menu Item: Resource ID": "",
            "Menu Item: Resource Handle": handle,
            "Menu Item: Collection Tags": "",
            "Menu Item: URL": f"/collections/{handle}" if handle else "",
            "Menu Item: Parent ID": parent_id_str,
            "Menu Item: Parent Title": parent_title,
            "Menu Item: Position": position_list[idx],
        })

    # ---- 11. 输出校验 ----
    if menu_rows:
        id_set = set()
        dup_ids = set()
        for row in menu_rows:
            mid = row.get("Menu Item: ID", "")
            if mid in id_set:
                dup_ids.add(mid)
            id_set.add(mid)
        if dup_ids:
            raise ValueError(f"生成菜单失败：存在重复 Menu Item ID - {', '.join(sorted(dup_ids))}")

        valid_ids = id_set | {"0", ""}
        invalid_parent_titles = []
        for row in menu_rows:
            pid = row.get("Menu Item: Parent ID", "0")
            if pid and pid not in valid_ids:
                invalid_parent_titles.append(row.get("Menu Item: Title", ""))
        if invalid_parent_titles:
            preview = ", ".join(invalid_parent_titles[:5])
            raise ValueError(f"生成菜单失败：存在无效父级 ID 的菜单项 - {preview}")

        blank_handle_titles = []
        for row in menu_rows:
            handle = row.get("Menu Item: Resource Handle", "").strip()
            if not handle:
                blank_handle_titles.append(row.get("Menu Item: Title", ""))
        if blank_handle_titles:
            preview = ", ".join(blank_handle_titles[:5])
            raise ValueError(f"生成菜单失败：存在空 Handle 的菜单项 - {preview}")

    return menu_rows



# ============================================================
# 工具函数：Pages → Menu（Policy / About 两个分组）
# ============================================================

def build_menu_from_pages(pages_dict, start_id, start_row_num):
    """根据 Pages 生成两个独立菜单：Policy 和 About（同参考逻辑）"""
    groups = [
        {
            "menu_id": "100000000002",
            "handle": "policy",
            "title": "Policy",
            "children": [
                "shipping-policy",
                "refund-policy",
                "privacy-policy",
                "customer-service-policy",
                "terms-of-purchase",
                "terms-of-use",
            ],
        },
        {
            "menu_id": "100000000003",
            "handle": "about",
            "title": "About",
            "children": [
                "about-us",
                "contact-us",
            ],
        },
    ]

    menu_rows = []
    current_id = start_id
    current_row = start_row_num

    for group in groups:
        group_pos = 0
        for page_key in group["children"]:
            page = pages_dict.get(page_key)
            if not page:
                continue
            handle = page["handle"]
            title = page["title"]
            group_pos += 1
            menu_rows.append({
                "ID": group["menu_id"],
                "Handle": group["handle"],
                "Command": "MERGE",
                "Title": group["title"],
                "Is Default": "TRUE",
                "Top Row": "TRUE" if group_pos == 1 else "",
                "Row #": str(current_row),
                "Menu Item: ID": str(current_id),
                "Menu Item: Title": title,
                "Menu Item: Command": "MERGE",
                "Menu Item: Resource Type": "PAGE",
                "Menu Item: Resource ID": "",
                "Menu Item: Resource Handle": handle,
                "Menu Item: Collection Tags": "",
                "Menu Item: URL": f"/pages/{handle}",
                "Menu Item: Parent ID": "",
                "Menu Item: Parent Title": "",
                "Menu Item: Position": str(group_pos),
            })
            current_id += 1
            current_row += 1

    return menu_rows


# ============================================================
# 产品 CSV 读取
# ============================================================

def read_product_csv(filepath):
    """读取产品 CSV，返回 (headers, rows)"""
    headers = []
    rows = []
    with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                headers = row
            else:
                if len(row) < len(headers):
                    row = row + [""] * (len(headers) - len(row))
                rows.append(row[:len(headers)])
    return headers, rows


# ============================================================
# Excel 写入工具
# ============================================================

def style_sheet(ws, headers, first_col_width=22):
    """给工作表设置表头样式、自动列宽、冻结首行（简约商务风）"""
    # 简约风格：浅灰底 + 深灰字 + 底部细线，避免厚重深蓝背景
    header_font = Font(bold=True, color="333333", size=11)
    header_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx, value=headers[col_idx - 1])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 自动列宽
    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for row in ws.iter_rows(min_row=2, max_col=col_idx, max_row=ws.max_row, min_col=col_idx):
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                line_len = max((len(line) for line in val.split("\n")), default=0)
                if line_len > max_len:
                    max_len = line_len
        width = min(max(max_len + 2, first_col_width if col_idx == 1 else 14), 80)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26


def write_dict_rows_to_sheet(wb, sheet_name, headers, rows_dicts):
    """写入字典格式的行到工作表。

    rows_dicts: list[dict]，每个 dict 的 key 对应 headers 中的列名
    """
    ws = wb.create_sheet(title=sheet_name)
    # 先写入表头（确保顺序正确）
    ws.append(headers)
    for row_dict in rows_dicts:
        ws.append([row_dict.get(col, "") for col in headers])
    style_sheet(ws, headers)
    return ws


def write_list_rows_to_sheet(wb, sheet_name, headers, rows):
    """写入 list[list] 格式的行到工作表"""
    ws = wb.create_sheet(title=sheet_name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws, headers)
    return ws


# ============================================================
# Altera 99 行限制：自动拆分写入工具
# ============================================================

# 每个 sheet 最多数据行数（Altera 限制 100，留 1 行余量）
ALTERA_MAX_ROWS_PER_SHEET = 99

# Excel 中保留的最大行数（超过此数的部分输出到 altera/ 文件夹的 CSV）
EXCEL_KEEP_MAX_ROWS = 99
# 每个 CSV 文件的最大行数
CSV_ROWS_PER_FILE = 99


def write_dict_rows_split_excel_csv(wb, base_sheet_name, headers, rows_dicts,
                                    csv_prefix, output_dir):
    """
    将行数据拆分写入 Excel + altera/CSV：
    - 前 EXCEL_KEEP_MAX_ROWS 行写入 Excel 的 base_sheet_name sheet
    - 超出部分每 CSV_ROWS_PER_FILE 行一个 CSV，输出到 output_dir/altera/
    - CSV 命名：{csv_prefix}_1.csv, {csv_prefix}_2.csv, ...

    返回: {
        "excel_sheets": [(sheet_name, row_count), ...],
        "csv_files": [(file_path, row_count), ...],
        "altera_dir": str,
    }
    """
    total = len(rows_dicts)
    result = {
        "excel_sheets": [],
        "csv_files": [],
        "altera_dir": None,
    }

    # ---- Excel 部分（前 EXCEL_KEEP_MAX_ROWS 行） ----
    excel_rows = rows_dicts[:EXCEL_KEEP_MAX_ROWS]
    if total <= EXCEL_KEEP_MAX_ROWS:
        # 全部放进 Excel，不需要 CSV
        ws = wb.create_sheet(title=base_sheet_name)
        ws.append(headers)
        for row_dict in excel_rows:
            ws.append([row_dict.get(col, "") for col in headers])
        style_sheet(ws, headers)
        result["excel_sheets"].append((base_sheet_name, len(excel_rows)))
        return result

    # 超过 99 行：前 99 行放 Excel
    ws = wb.create_sheet(title=base_sheet_name)
    ws.append(headers)
    for row_dict in excel_rows:
        ws.append([row_dict.get(col, "") for col in headers])
    style_sheet(ws, headers)
    result["excel_sheets"].append((base_sheet_name, len(excel_rows)))

    # ---- CSV 部分（超出 99 行的部分） ----
    altera_dir = os.path.join(output_dir, "altera")
    os.makedirs(altera_dir, exist_ok=True)
    result["altera_dir"] = altera_dir

    remaining = rows_dicts[EXCEL_KEEP_MAX_ROWS:]
    file_idx = 1
    for start in range(0, len(remaining), CSV_ROWS_PER_FILE):
        end = min(start + CSV_ROWS_PER_FILE, len(remaining))
        chunk = remaining[start:end]
        csv_filename = f"{csv_prefix}_{file_idx}.csv"
        csv_path = os.path.join(altera_dir, csv_filename)

        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row_dict in chunk:
                writer.writerow([row_dict.get(col, "") for col in headers])

        result["csv_files"].append((csv_path, len(chunk)))
        file_idx += 1

    return result


def write_dict_rows_to_csv_only(rows_dicts, headers, csv_prefix, output_dir):
    """
    将所有行数据输出到 CSV 文件（不写入 Excel）。
    每 CSV_ROWS_PER_FILE 行一个 CSV，输出到 output_dir/altera/
    CSV 命名：{csv_prefix}_1.csv, {csv_prefix}_2.csv, ...

    返回: {
        "csv_files": [(file_path, row_count), ...],
        "altera_dir": str,
    }
    """
    if not rows_dicts:
        return {"csv_files": [], "altera_dir": None}

    altera_dir = os.path.join(output_dir, "altera")
    os.makedirs(altera_dir, exist_ok=True)

    csv_files = []
    file_idx = 1
    for start in range(0, len(rows_dicts), CSV_ROWS_PER_FILE):
        end = min(start + CSV_ROWS_PER_FILE, len(rows_dicts))
        chunk = rows_dicts[start:end]
        csv_filename = f"{csv_prefix}_{file_idx}.csv"
        csv_path = os.path.join(altera_dir, csv_filename)

        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row_dict in chunk:
                writer.writerow([row_dict.get(col, "") for col in headers])

        csv_files.append((csv_path, len(chunk)))
        file_idx += 1

    return {"csv_files": csv_files, "altera_dir": altera_dir}


def write_dict_rows_auto_split(wb, base_sheet_name, headers, rows_dicts):
    """
    写入字典行，超过 ALTERA_MAX_ROWS_PER_SHEET 行时自动拆分为多个 sheet。
    sheet 名规则：
      - 若 rows <= 99:  SheetName
      - 若 rows > 99:   SheetName_1, SheetName_2, ...
    返回 list[(sheet_name, row_count)] 写入信息列表。
    """
    total = len(rows_dicts)
    sheet_infos = []

    if total == 0:
        # 无数据也至少建一个空表（只有表头）
        ws = wb.create_sheet(title=base_sheet_name)
        ws.append(headers)
        style_sheet(ws, headers)
        sheet_infos.append((base_sheet_name, 0))
        return sheet_infos

    if total <= ALTERA_MAX_ROWS_PER_SHEET:
        # 不需要拆分
        ws = wb.create_sheet(title=base_sheet_name)
        ws.append(headers)
        for row_dict in rows_dicts:
            ws.append([row_dict.get(col, "") for col in headers])
        style_sheet(ws, headers)
        sheet_infos.append((base_sheet_name, total))
        return sheet_infos

    # 需要拆分
    part_idx = 1
    for start in range(0, total, ALTERA_MAX_ROWS_PER_SHEET):
        end = min(start + ALTERA_MAX_ROWS_PER_SHEET, total)
        chunk = rows_dicts[start:end]
        sheet_name = f"{base_sheet_name}_{part_idx}"
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        for row_dict in chunk:
            ws.append([row_dict.get(col, "") for col in headers])
        style_sheet(ws, headers)
        sheet_infos.append((sheet_name, len(chunk)))
        part_idx += 1

    return sheet_infos


def write_list_rows_auto_split(wb, base_sheet_name, headers, rows):
    """同 write_dict_rows_auto_split，但 rows 为 list[list]。"""
    total = len(rows)
    sheet_infos = []

    if total == 0:
        ws = wb.create_sheet(title=base_sheet_name)
        ws.append(headers)
        style_sheet(ws, headers)
        sheet_infos.append((base_sheet_name, 0))
        return sheet_infos

    if total <= ALTERA_MAX_ROWS_PER_SHEET:
        ws = wb.create_sheet(title=base_sheet_name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        style_sheet(ws, headers)
        sheet_infos.append((base_sheet_name, total))
        return sheet_infos

    part_idx = 1
    for start in range(0, total, ALTERA_MAX_ROWS_PER_SHEET):
        end = min(start + ALTERA_MAX_ROWS_PER_SHEET, total)
        chunk = rows[start:end]
        sheet_name = f"{base_sheet_name}_{part_idx}"
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        for row in chunk:
            ws.append(row)
        style_sheet(ws, headers)
        sheet_infos.append((sheet_name, len(chunk)))
        part_idx += 1

    return sheet_infos


# ============================================================
# 主生成函数：把一切从 CSV 生成
# ============================================================

def generate_excel(
    domain,
    brand,
    product_csv_path=None,
    output_path=None,
    with_domain=False,
):
    """根据产品 CSV 数据，生成 Altera 导入用的 Excel 文件。

    完全基于 CSV 产品数据生成 SEO title/description 和页面内容，
    不再依赖 title/description 输入。

    with_domain: False=不带域名版本（纯文本），True=带域名版本（可点击链接）

    核心逻辑（4 个工作表，不含 Products 原始表）:
      1. Metafield Definitions ← 产品 CSV 的 c_f / product.metafields 列
      2. Smart Collections ← 产品 CSV 的 Tags 列（每个 unique tag 1 条）
      3. Pages ← 从 CSV 产品关键词生成 SEO 内容 + 页面模板
      4. Menu ← 商品分类（来自 Smart Collections）+ Policy / About 两个分组（来自 Pages）
    """
    # 参数校验
    if not product_csv_path or not os.path.exists(product_csv_path):
        raise FileNotFoundError(f"找不到产品 CSV 文件: {product_csv_path}")

    # 1. 从 CSV 动态生成内容
    metafield_rows = build_metafield_definitions_from_csv(product_csv_path)
    collection_rows = build_smart_collections_from_csv(product_csv_path)

    # 1.0.1 热销关键词统计（从 Title 列提取）
    hot_keywords = extract_hot_keywords_from_titles(product_csv_path, top_n=15)

    # 1.0.1 注意：按参考代码，**不在 collection 阶段生成前缀节点**。
    # 中间层级仅在菜单阶段通过 infer_missing_group_titles 严格推断（满足
    # 单词数 > 2、必须已有上层父级、子项≥2 等约束）。inferred 菜单节点
    # 的 Handle 会复用父级 collection 的 Handle（find_parent_handle_for_inferred_title），
    # 所以不会报 MENU009，也不需要额外的 prefix collections。
    #
    # 为什么不在此阶段生成前缀 collections？
    # - 会产生大量无意义的"Best"/"Shop"/"Freezer"等单字集合
    # - 违反参考代码逻辑（参考代码仅在菜单层推断，不在 collection 层推断）
    # - 导入时可能出现空行/额外噪点

    # 仅重新编号 Row #
    for idx, col in enumerate(collection_rows):
        col["Row #"] = idx + 1  # int 而非 str（Excel 中以数字形式显示）

    # 1.1 菜单来自两部分：(a) Smart Collections（商品分类）, (b) Pages（Policy / About 分组）
    # 注意：传给 build_menu_from_collections 的是**所有 collection 行**，
    # 但菜单生成时只以 "有 Title 的行"（即每个 collection 的头行）为准。
    collection_menu_rows = build_menu_from_collections(collection_rows)

    # 2. 生成 Pages（完全基于 CSV 产品关键词生成内容和 SEO）
    # with_domain: False=不带域名版本，True=带域名版本
    content = generate_all_content(
        domain, brand=brand, product_csv_path=product_csv_path,
        with_domain=with_domain
    )
    pages_rows = []
    for key, page in content["pages"].items():
        pages_rows.append({
            "Handle": page["handle"],
            "Title": page["title"],
            "Body (HTML)": page["body_html"],
            "Template Suffix": "",
            "Published": "TRUE",
        })

    # 2.1 从 Pages 构建 Policy / About 分组菜单，追加到 collection 菜单之后
    next_id = len(collection_menu_rows) + 1
    next_row = len(collection_menu_rows) + 1
    pages_menu_rows = build_menu_from_pages(content["pages"], next_id, next_row)
    menu_rows = collection_menu_rows + pages_menu_rows

    # 3. 创建工作簿并写入（跳过 Products 原始表，只保留 4 个工作表）
    wb = Workbook()
    # 删除默认创建的 Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 用于收集每个 sheet 的写入/拆分信息，用于输出提示
    all_sheet_infos = []
    # 收集 CSV 文件信息（Smart Collections 和 Menu 超出 99 行的部分）
    csv_files_info = []

    # 4.0 计算输出目录（用于 altera/CSV 输出，以及版本分文件夹）
    csv_dir = os.path.dirname(os.path.abspath(product_csv_path))
    safe_brand = "".join(c for c in brand if c.isalnum() or c in "-_")[:40] or "store"

    if not output_path:
        # 根据版本分文件夹：带域名版本用 _with_domain/，不带域名版本用 _no_domain/
        version_folder = "_with_domain" if with_domain else "_no_domain"
        folder_path = os.path.join(csv_dir, safe_brand + version_folder)
        os.makedirs(folder_path, exist_ok=True)
        output_path = os.path.join(
            folder_path,
            f"{safe_brand}_shopify_altera.xlsx"
        )
    output_dir = os.path.dirname(os.path.abspath(output_path))

    # 4.1 Metafield Definitions（每 99 行自动拆分）
    meta_infos = write_dict_rows_auto_split(wb, "Metafield Definitions",
                                             METAFIELD_DEFINITION_COLUMNS, metafield_rows)
    all_sheet_infos.append(("Metafield Definitions", len(metafield_rows), meta_infos))

    # 4.2 判断 Smart Collections 是否超过 99 行
    # 超过 99 行：不在 Excel 中创建工作表，所有数据输出到 CSV
    # 未超过 99 行：在 Excel 中创建工作表
    sc_exceeds_limit = len(collection_rows) > EXCEL_KEEP_MAX_ROWS
    if sc_exceeds_limit:
        # 全部输出到 CSV
        sc_result = write_dict_rows_to_csv_only(
            collection_rows, SMART_COLLECTION_COLUMNS,
            csv_prefix="collection", output_dir=output_dir
        )
        csv_files_info.append(("Smart Collections", sc_result["csv_files"]))
    else:
        # 写入 Excel
        ws = wb.create_sheet(title="Smart Collections")
        ws.append(SMART_COLLECTION_COLUMNS)
        for row_dict in collection_rows:
            ws.append([row_dict.get(col, "") for col in SMART_COLLECTION_COLUMNS])
        style_sheet(ws, SMART_COLLECTION_COLUMNS)
        all_sheet_infos.append(("Smart Collections", len(collection_rows), [("Smart Collections", len(collection_rows))]))

    # 4.3 Pages（每 99 行自动拆分）
    pages_infos = write_dict_rows_auto_split(wb, "Pages", PAGES_COLUMNS, pages_rows)
    all_sheet_infos.append(("Pages", len(pages_rows), pages_infos))

    # 4.4 判断 Menu 是否超过 99 行
    menu_exceeds_limit = len(menu_rows) > EXCEL_KEEP_MAX_ROWS
    if menu_exceeds_limit:
        # 全部输出到 CSV
        menu_result = write_dict_rows_to_csv_only(
            menu_rows, MENU_COLUMNS,
            csv_prefix="menu", output_dir=output_dir
        )
        csv_files_info.append(("Menu", menu_result["csv_files"]))
    else:
        # 写入 Excel
        ws = wb.create_sheet(title="Menu")
        ws.append(MENU_COLUMNS)
        for row_dict in menu_rows:
            ws.append([row_dict.get(col, "") for col in MENU_COLUMNS])
        style_sheet(ws, MENU_COLUMNS)
        all_sheet_infos.append(("Menu", len(menu_rows), [("Menu", len(menu_rows))]))

    wb.save(output_path)

    # 5. 生成定制化主题 zip 包（含自定义 tabs + 热搜词）
    theme_zip_path = build_theme_zip(brand, domain, metafield_rows, output_dir, hot_keywords=hot_keywords)

    # 6. 产品 CSV 大于 14MB 时自动分割
    split_csv_files = []
    product_csv_size_mb = os.path.getsize(product_csv_path) / (1024 * 1024)
    if product_csv_size_mb > 14:
        split_csv_files = split_product_csv(product_csv_path, brand, output_dir)

    summary_lines = [
        f"✓ 已生成文件: {os.path.basename(output_path)}",
        f"  · 源文件: {os.path.basename(product_csv_path)}",
        f"  · 工作表总览:",
    ]
    for base_name, total_rows, sheet_infos in all_sheet_infos:
        if len(sheet_infos) > 1:
            part_detail = ", ".join(f"{name}={rows}行" for name, rows in sheet_infos)
            summary_lines.append(f"      → {base_name}: {total_rows} 行 → 已拆分为 {len(sheet_infos)} 个表 ({part_detail})")
        else:
            summary_lines.append(f"      → {base_name}: {total_rows} 行（无需拆分）")

    if csv_files_info:
        summary_lines.append("")
        summary_lines.append(f"  · 超出 99 行的部分已输出到 altera/ 文件夹（CSV 格式）:")
        for base_name, files in csv_files_info:
            for fpath, frows in files:
                summary_lines.append(f"      → {os.path.basename(fpath)}: {frows} 行")

    summary_lines.append("")
    summary_lines.append(f"✓ 已生成主题包: {os.path.basename(theme_zip_path)}")
    summary_lines.append(f"  · 自定义 Tabs: {len(metafield_rows)} 个")

    if split_csv_files:
        summary_lines.append("")
        summary_lines.append(f"✓ 产品 CSV 超过 14MB，已自动分割为 {len(split_csv_files)} 个文件:")
        for fp in split_csv_files:
            size_mb = os.path.getsize(fp) / (1024 * 1024)
            summary_lines.append(f"      → {os.path.basename(fp)} ({size_mb:.1f} MB)")

    summary_text = "\n".join(summary_lines)

    return {
        "output_path": output_path,
        "theme_zip_path": theme_zip_path,
        "split_csv_files": split_csv_files,
        "summary": summary_text,
        "metafield_count": len(metafield_rows),
        "collection_count": len(collection_rows),
        "menu_count": len(menu_rows),
        "pages_count": len(pages_rows),
        "context": content["context"],
        "seo_title": content["seo_title"],
        "seo_description": content["seo_description"],
        "menu_rows": menu_rows,
        "all_sheet_infos": all_sheet_infos,
        "csv_files_info": csv_files_info,
        "hot_keywords": hot_keywords,
    }


# ============================================================
# GUI 部分
# ============================================================

class AlteraExcelGeneratorApp:
    def __init__(self, root):
        self.root = root
        root.title("豆脚Altera Excel 生成工具")
        root.geometry("1240x820")
        root.minsize(1100, 720)

        # 设置窗口图标（优先 favicon.png，兼容 icon.ico 作为备用）
        # 兼容 PyInstaller 打包（从 exe 所在目录读取）和源码运行两种模式
        icon_candidates = [
            # 优先 favicon.png（与 exe 嵌入图标保持一致）
            os.path.join(RESOURCE_PATH, "favicon.png"),
            os.path.join(RESOURCE_PATH, "data", "favicon.png"),
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "favicon.png"),
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "favicon.png"),
            # 备用：icon.ico
            os.path.join(RESOURCE_PATH, "icon.ico"),
            os.path.join(RESOURCE_PATH, "data", "icon.ico"),
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "icon.ico"),
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "icon.ico"),
        ]
        for ic_path in icon_candidates:
            if os.path.exists(ic_path):
                try:
                    if ic_path.lower().endswith(".png"):
                        # PNG 图标：用 PhotoImage + wm_iconbitmap 的组合方式
                        img = tk.PhotoImage(file=ic_path)
                        root.tk.call("wm", "iconphoto", root._w, img)
                        self._app_icon_image = img  # 保持引用防止 GC
                        self._app_icon_path = ic_path
                    else:
                        # ICO 图标：直接用 iconbitmap
                        root.iconbitmap(ic_path)
                        self._app_icon_path = ic_path  # 保持引用防止 GC
                    break
                except Exception:
                    continue

        # Design tokens
        self.brand_color = "#0B57D0"
        self.brand_hover = "#1967E0"
        self.brand_press = "#0A4AAF"
        self.surface = "#FAFBFC"
        self.surface_alt = "#F1F3F5"
        self.text = "#1A1A1A"
        self.text_mid = "#5E6A75"
        self.text_soft = "#8A96A3"
        self.line = "#E5E8EB"
        self.field_bg = "#FFFFFF"
        self.field_border = "#D4D9E0"
        self.output_bg = "#0E1116"
        self.output_text = "#D7E0E8"
        self.success_color = "#0F9960"

        root.configure(bg=self.surface)

        app = tk.Frame(root, bg=self.surface)
        app.pack(fill="both", expand=True)

        # ===== TOP BAR =====
        bar = tk.Frame(app, bg="#FFFFFF", highlightbackground=self.line, highlightthickness=1)
        bar.pack(fill="x")

        bar_inner = tk.Frame(bar, bg="#FFFFFF")
        bar_inner.pack(fill="x", padx=32, pady=18)

        tk.Label(bar_inner, text="豆脚Altera Excel 生成工具",
                 bg="#FFFFFF", fg=self.text,
                 font=("Microsoft YaHei UI", 15, "bold")).pack(side="left")

        # 右上角：版本号 + 检查更新按钮
        right_wrap = tk.Frame(bar_inner, bg="#FFFFFF")
        right_wrap.pack(side="right")

        self.version_label = tk.Label(right_wrap, text=f"v{APP_VERSION}",
                                      bg="#FFFFFF", fg=self.text_soft,
                                      font=("Microsoft YaHei UI", 10))
        self.version_label.pack(side="left", padx=(0, 12))

        self.check_update_btn = tk.Button(
            right_wrap, text="检查更新",
            bg="#FFFFFF", fg=self.brand_color,
            activebackground=self.surface_alt, activeforeground=self.brand_press,
            relief="flat", bd=0, cursor="hand2",
            font=("Microsoft YaHei UI", 10),
            command=self._on_check_update
        )
        self.check_update_btn.pack(side="right")

        # ===== TAB BAR =====
        tab_bar = tk.Frame(app, bg="#FFFFFF", highlightbackground=self.line, highlightthickness=1)
        tab_bar.pack(fill="x")

        tab_inner = tk.Frame(tab_bar, bg="#FFFFFF", height=48)
        tab_inner.pack(fill="x", padx=32)
        tab_inner.pack_propagate(False)

        self.current_tab = "excel"

        self.excel_tab = tk.Frame(tab_inner, bg="#FFFFFF")
        self.excel_tab.pack(side="left", fill="y")

        self.utils_tab = tk.Frame(tab_inner, bg="#FFFFFF")
        self.utils_tab.pack(side="left", fill="y")

        self.terms_tab = tk.Frame(tab_inner, bg="#FFFFFF")
        self.terms_tab.pack(side="left", fill="y")

        self.excel_tab_label = tk.Label(self.excel_tab, text="文件处理",
                                        bg="#FFFFFF", fg=self.brand_color,
                                        font=("Microsoft YaHei UI", 10, "bold"),
                                        cursor="hand2")
        self.excel_tab_label.pack(expand=True, padx=20)

        self.utils_tab_label = tk.Label(self.utils_tab, text="复制站模板处理",
                                        bg="#FFFFFF", fg=self.text_mid,
                                        font=("Microsoft YaHei UI", 10),
                                        cursor="hand2")
        self.utils_tab_label.pack(expand=True, padx=20)

        self.terms_tab_label = tk.Label(self.terms_tab, text="条款页面生成",
                                        bg="#FFFFFF", fg=self.text_mid,
                                        font=("Microsoft YaHei UI", 10),
                                        cursor="hand2")
        self.terms_tab_label.pack(expand=True, padx=20)

        # Tab indicator
        self.excel_indicator = tk.Frame(tab_inner, bg=self.brand_color, height=2)
        self.excel_indicator.place(x=0, y=46, width=80)

        def _update_indicator():
            if self.current_tab == "excel":
                x = self.excel_tab.winfo_x()
                w = self.excel_tab.winfo_width()
                if w > 0:
                    self.excel_indicator.place(x=x, y=46, width=w)
            elif self.current_tab == "utils":
                x = self.utils_tab.winfo_x()
                w = self.utils_tab.winfo_width()
                if w > 0:
                    self.excel_indicator.place(x=x, y=46, width=w)
            else:
                x = self.terms_tab.winfo_x()
                w = self.terms_tab.winfo_width()
                if w > 0:
                    self.excel_indicator.place(x=x, y=46, width=w)

        def _ensure_indicator():
            _update_indicator()
            tab_inner.after(50, _update_indicator)
            tab_inner.after(200, _update_indicator)
            tab_inner.after(500, _update_indicator)

        tab_inner.after(10, _ensure_indicator)

        def switch_to_excel(e=None):
            if self.current_tab == "excel":
                return
            self.current_tab = "excel"
            self.excel_tab_label.configure(fg=self.brand_color, font=("Microsoft YaHei UI", 10, "bold"))
            self.utils_tab_label.configure(fg=self.text_mid, font=("Microsoft YaHei UI", 10))
            self.terms_tab_label.configure(fg=self.text_mid, font=("Microsoft YaHei UI", 10))
            _update_indicator()
            self.excel_workspace.pack(fill="both", expand=True, padx=32, pady=24)
            self.utils_workspace.pack_forget()
            self.terms_workspace.pack_forget()

        def switch_to_utils(e=None):
            if self.current_tab == "utils":
                return
            self.current_tab = "utils"
            self.utils_tab_label.configure(fg=self.brand_color, font=("Microsoft YaHei UI", 10, "bold"))
            self.excel_tab_label.configure(fg=self.text_mid, font=("Microsoft YaHei UI", 10))
            self.terms_tab_label.configure(fg=self.text_mid, font=("Microsoft YaHei UI", 10))
            _update_indicator()
            self.utils_workspace.pack(fill="both", expand=True, padx=32, pady=24)
            self.excel_workspace.pack_forget()
            self.terms_workspace.pack_forget()

        def switch_to_terms(e=None):
            if self.current_tab == "terms":
                return
            self.current_tab = "terms"
            self.terms_tab_label.configure(fg=self.brand_color, font=("Microsoft YaHei UI", 10, "bold"))
            self.excel_tab_label.configure(fg=self.text_mid, font=("Microsoft YaHei UI", 10))
            self.utils_tab_label.configure(fg=self.text_mid, font=("Microsoft YaHei UI", 10))
            _update_indicator()
            self.terms_workspace.pack(fill="both", expand=True, padx=32, pady=24)
            self.excel_workspace.pack_forget()
            self.utils_workspace.pack_forget()

        self.excel_tab.bind("<Button-1>", switch_to_excel)
        self.excel_tab_label.bind("<Button-1>", switch_to_excel)
        self.utils_tab.bind("<Button-1>", switch_to_utils)
        self.utils_tab_label.bind("<Button-1>", switch_to_utils)
        self.terms_tab.bind("<Button-1>", switch_to_terms)
        self.terms_tab_label.bind("<Button-1>", switch_to_terms)

        def _on_tab_enter_excel(e):
            if self.current_tab != "excel":
                self.excel_tab_label.configure(fg=self.text)
        def _on_tab_leave_excel(e):
            if self.current_tab != "excel":
                self.excel_tab_label.configure(fg=self.text_mid)
        def _on_tab_enter_utils(e):
            if self.current_tab != "utils":
                self.utils_tab_label.configure(fg=self.text)
        def _on_tab_leave_utils(e):
            if self.current_tab != "utils":
                self.utils_tab_label.configure(fg=self.text_mid)
        def _on_tab_enter_terms(e):
            if self.current_tab != "terms":
                self.terms_tab_label.configure(fg=self.text)
        def _on_tab_leave_terms(e):
            if self.current_tab != "terms":
                self.terms_tab_label.configure(fg=self.text_mid)

        self.excel_tab.bind("<Enter>", _on_tab_enter_excel)
        self.excel_tab_label.bind("<Enter>", _on_tab_enter_excel)
        self.excel_tab.bind("<Leave>", _on_tab_leave_excel)
        self.excel_tab_label.bind("<Leave>", _on_tab_leave_excel)
        self.utils_tab.bind("<Enter>", _on_tab_enter_utils)
        self.utils_tab_label.bind("<Enter>", _on_tab_enter_utils)
        self.utils_tab.bind("<Leave>", _on_tab_leave_utils)
        self.utils_tab_label.bind("<Leave>", _on_tab_leave_utils)
        self.terms_tab.bind("<Enter>", _on_tab_enter_terms)
        self.terms_tab_label.bind("<Enter>", _on_tab_enter_terms)
        self.terms_tab.bind("<Leave>", _on_tab_leave_terms)
        self.terms_tab_label.bind("<Leave>", _on_tab_leave_terms)

        # ===== EXCEL WORKSPACE =====
        self.excel_workspace = tk.Frame(app, bg=self.surface)
        self.excel_workspace.pack(fill="both", expand=True, padx=32, pady=24)

        # ---- LEFT: SCROLLABLE INPUT COLUMN ----
        left_outer = tk.Frame(self.excel_workspace, bg=self.surface)
        left_outer.pack(side="left", fill="y")

        # Canvas for scrollable content area
        self._left_canvas = tk.Canvas(left_outer, width=440, bg=self.surface,
                                       highlightthickness=0, bd=0)
        self._left_canvas.pack(side="left", fill="y", expand=False)

        left_scroll = ttk.Scrollbar(left_outer, orient="vertical",
                                     command=self._left_canvas.yview)
        left_scroll.pack(side="left", fill="y")
        self._left_canvas.configure(yscrollcommand=left_scroll.set)

        left = tk.Frame(self._left_canvas, bg=self.surface)
        self._left_window = self._left_canvas.create_window(
            (0, 0), window=left, anchor="nw", width=440
        )

        def _on_left_configure(event):
            bbox = self._left_canvas.bbox("all")
            if bbox:
                self._left_canvas.configure(scrollregion=bbox)

        left.bind("<Configure>", _on_left_configure)
        self._left_canvas.bind("<Configure>",
                               lambda e: self._left_canvas.itemconfigure(
                                   self._left_window, width=e.width))

        # Mouse wheel scroll on left panel
        def _on_mousewheel(event):
            self._left_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _bind_mwheel(widget):
            widget.bind("<MouseWheel>", _on_mousewheel)
            for child in widget.winfo_children():
                _bind_mwheel(child)

        _bind_mwheel(left)
        _bind_mwheel(self._left_canvas)

        # ---- CSV 文件 ----
        csv_wrap = tk.Frame(left, bg=self.surface)
        csv_wrap.pack(fill="x", pady=(0, 18))

        tk.Label(csv_wrap, text="产品文件路径",
                 bg=self.surface, fg=self.text,
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w")

        path_row = tk.Frame(csv_wrap, bg=self.surface)
        path_row.pack(fill="x", pady=(6, 0))

        self.csv_path_var = tk.StringVar()
        self.brand_var = tk.StringVar()
        default_csv = self._auto_find_shopify_csv()
        if default_csv:
            self.csv_path_var.set(default_csv)
            # 自动识别品牌名
            filename = os.path.splitext(os.path.basename(default_csv))[0]
            parts = filename.split("_")
            if parts and parts[0] and not parts[0].isdigit():
                self.brand_var.set(parts[0].capitalize())

        self.csv_entry = tk.Entry(path_row, textvariable=self.csv_path_var,
                                  bg=self.field_bg, fg=self.text,
                                  relief="flat", insertwidth=2,
                                  insertbackground=self.brand_color,
                                  highlightthickness=1,
                                  highlightbackground=self.field_border,
                                  highlightcolor=self.brand_color,
                                  font=("Microsoft YaHei UI", 10))
        self.csv_entry.pack(side="left", fill="x", expand=True, ipady=8, ipadx=10)

        browse_btn = tk.Button(path_row, text="浏览",
                               command=self._browse_csv,
                               bg=self.surface_alt, fg=self.text,
                               activebackground=self.line,
                               activeforeground=self.text,
                               relief="flat", font=("Microsoft YaHei UI", 9, "bold"),
                               cursor="hand2", padx=16, pady=8, bd=0,
                               highlightthickness=0)
        browse_btn.pack(side="left", padx=(10, 0))
        self._bind_hover(browse_btn, self.surface_alt, self.line)

        # ---- 品牌信息 ----
        form = tk.Frame(left, bg=self.surface)
        form.pack(fill="x", pady=(0, 18))

        self._labelled_field(form, "品牌", self.brand_var, hint="")[0].pack(fill="x", pady=(0, 12))

        self.domain_var = tk.StringVar()
        domain_wrap, self.domain_entry = self._labelled_field(form, "域名", self.domain_var, hint="")
        domain_wrap.pack(fill="x", pady=(12, 0))

        # ---- 条款版本（分段切换） ----
        version_wrap = tk.Frame(left, bg=self.surface)
        version_wrap.pack(fill="x", pady=(0, 18))

        tk.Label(version_wrap, text="条款版本",
                 bg=self.surface, fg=self.text,
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w", pady=(0, 8))

        self.page_version_var = tk.StringVar(value="with_domain")

        seg_container = tk.Frame(version_wrap, bg=self.surface_alt,
                                  highlightthickness=0, bd=0)
        seg_container.pack(fill="x")

        seg_left = tk.Frame(seg_container, bg=self.brand_color, height=40)
        seg_left.pack(side="left", fill="both", expand=True)
        seg_left.pack_propagate(False)

        seg_right = tk.Frame(seg_container, bg=self.surface_alt, height=40)
        seg_right.pack(side="left", fill="both", expand=True)
        seg_right.pack_propagate(False)

        left_label = tk.Label(seg_left, text="带域名",
                               bg=self.brand_color, fg="#FFFFFF",
                               font=("Microsoft YaHei UI", 9, "bold"),
                               cursor="hand2")
        left_label.pack(expand=True)

        right_label = tk.Label(seg_right, text="不带域名",
                                bg=self.surface_alt, fg=self.text_mid,
                                font=("Microsoft YaHei UI", 9),
                                cursor="hand2")
        right_label.pack(expand=True)

        hint_label = tk.Label(version_wrap, text="条款含品牌名和可点击域名链接",
                               bg=self.surface, fg=self.text_soft,
                               font=("Microsoft YaHei UI", 8))
        hint_label.pack(anchor="w", pady=(6, 0))

        def on_version_change():
            if self.page_version_var.get() == "with_domain":
                seg_left.configure(bg=self.brand_color)
                left_label.configure(bg=self.brand_color, fg="#FFFFFF", font=("Microsoft YaHei UI", 9, "bold"))
                seg_right.configure(bg=self.surface_alt)
                right_label.configure(bg=self.surface_alt, fg=self.text_mid, font=("Microsoft YaHei UI", 9))
                hint_label.configure(text="条款含品牌名和可点击域名链接")
                self.domain_entry.configure(state="normal", disabledforeground=self.text)
            else:
                seg_left.configure(bg=self.surface_alt)
                left_label.configure(bg=self.surface_alt, fg=self.text_mid, font=("Microsoft YaHei UI", 9))
                seg_right.configure(bg=self.brand_color)
                right_label.configure(bg=self.brand_color, fg="#FFFFFF", font=("Microsoft YaHei UI", 9, "bold"))
                hint_label.configure(text="条款仅含品牌名，无域名无邮箱")
                self.domain_entry.configure(state="disabled", disabledbackground=self.surface_alt, disabledforeground=self.text_soft)

        def select_with_domain(e=None):
            if self.page_version_var.get() != "with_domain":
                self.page_version_var.set("with_domain")
                on_version_change()

        def select_no_domain(e=None):
            if self.page_version_var.get() != "no_domain":
                self.page_version_var.set("no_domain")
                on_version_change()

        left_label.bind("<Button-1>", select_with_domain)
        seg_left.bind("<Button-1>", select_with_domain)
        right_label.bind("<Button-1>", select_no_domain)
        seg_right.bind("<Button-1>", select_no_domain)

        def _on_enter_left(e):
            if self.page_version_var.get() != "with_domain":
                seg_left.configure(bg=self.line)
                left_label.configure(bg=self.line, fg=self.text)
        def _on_leave_left(e):
            if self.page_version_var.get() != "with_domain":
                seg_left.configure(bg=self.surface_alt)
                left_label.configure(bg=self.surface_alt, fg=self.text_mid)
        def _on_enter_right(e):
            if self.page_version_var.get() != "no_domain":
                seg_right.configure(bg=self.line)
                right_label.configure(bg=self.line, fg=self.text)
        def _on_leave_right(e):
            if self.page_version_var.get() != "no_domain":
                seg_right.configure(bg=self.surface_alt)
                right_label.configure(bg=self.surface_alt, fg=self.text_mid)

        left_label.bind("<Enter>", _on_enter_left)
        seg_left.bind("<Enter>", _on_enter_left)
        left_label.bind("<Leave>", _on_leave_left)
        seg_left.bind("<Leave>", _on_leave_left)
        right_label.bind("<Enter>", _on_enter_right)
        seg_right.bind("<Enter>", _on_enter_right)
        right_label.bind("<Leave>", _on_leave_right)
        seg_right.bind("<Leave>", _on_leave_right)

        # ---- 生成按钮 ----
        self.generate_btn = tk.Button(left, text="开始处理",
                                      command=self.on_generate,
                                      bg=self.brand_color, fg="#FFFFFF",
                                      activebackground=self.brand_press,
                                      activeforeground="#FFFFFF",
                                      relief="flat", font=("Microsoft YaHei UI", 10, "bold"),
                                      cursor="hand2", padx=24, pady=12, bd=0,
                                      highlightthickness=0)
        self.generate_btn.pack(fill="x", ipady=4, pady=(8, 8))
        self._bind_hover(self.generate_btn, self.brand_color, self.brand_hover)

        # ---- 次要操作 ----
        secondary = tk.Frame(left, bg=self.surface)
        secondary.pack(fill="x")
        of_btn = tk.Button(secondary, text="打开文件",
                          command=self.on_open_file,
                          bg=self.surface, fg=self.text_mid,
                          activebackground=self.surface, activeforeground=self.brand_color,
                          relief="flat", font=("Microsoft YaHei UI", 9),
                          cursor="hand2", bd=0, highlightthickness=0)
        of_btn.pack(side="left")
        self._bind_text_hover(of_btn, self.text_mid, self.brand_color)

        tk.Label(secondary, text="\u00b7", bg=self.surface, fg=self.text_soft,
                 font=("Microsoft YaHei UI", 9)).pack(side="left", padx=10)

        od_btn = tk.Button(secondary, text="打开文件夹",
                          command=self.on_open_folder,
                          bg=self.surface, fg=self.text_mid,
                          activebackground=self.surface, activeforeground=self.brand_color,
                          relief="flat", font=("Microsoft YaHei UI", 9),
                          cursor="hand2", bd=0, highlightthickness=0)
        od_btn.pack(side="left")
        self._bind_text_hover(od_btn, self.text_mid, self.brand_color)

        # RIGHT: OUTPUT COLUMN
        right = tk.Frame(self.excel_workspace, bg=self.surface)
        right.pack(side="left", fill="both", expand=True, padx=(24, 0))

        tk.Label(right, text="输出", bg=self.surface, fg=self.text,
                 font=("Microsoft YaHei UI", 11, "bold")).pack(anchor="w")

        out_container = tk.Frame(right, bg=self.output_bg, highlightthickness=0)
        out_container.pack(fill="both", expand=True, pady=(8, 0))

        self.result_text = tk.Text(out_container, wrap="word",
                                    bg=self.output_bg, fg=self.output_text,
                                    relief="flat", font=("Consolas", 11),
                                    padx=20, pady=16,
                                    insertbackground=self.output_text,
                                    highlightthickness=0,
                                    selectbackground="#2A3441",
                                    selectforeground="#FFFFFF",
                                    cursor="xterm")
        self.result_text.pack(fill="both", expand=True)

        self.result_text.tag_configure("muted", foreground="#6D7B8A", font=("Consolas", 11))
        self.result_text.tag_configure("accent", foreground="#4CB3FF", font=("Consolas", 11, "bold"))
        self.result_text.tag_configure("ok", foreground="#5FD7A2", font=("Consolas", 11, "bold"))
        self.result_text.tag_configure("warn", foreground="#FFB454", font=("Consolas", 11, "bold"))
        self.result_text.tag_configure("err", foreground="#FF7A7A", font=("Consolas", 11, "bold"))
        self.result_text.tag_configure("line", foreground="#3A4653", font=("Consolas", 11))
        self.result_text.tag_configure("key", foreground="#D4A5FF", font=("Consolas", 11))

        self._write_initial()
        self.result_text.configure(state="disabled")

        # ===== UTILS WORKSPACE =====
        self.utils_workspace = tk.Frame(app, bg=self.surface)

        utils_cols = tk.Frame(self.utils_workspace, bg=self.surface)
        utils_cols.pack(fill="both", expand=True)

        utils_left_col = tk.Frame(utils_cols, bg=self.surface, width=460)
        utils_left_col.pack(side="left", fill="y", padx=(0, 24))
        utils_left_col.pack_propagate(False)

        utils_right_col = tk.Frame(utils_cols, bg=self.surface, width=460)
        utils_right_col.pack(side="left", fill="y")
        utils_right_col.pack_propagate(False)

        # 左列
        utils_form_left = tk.Frame(utils_left_col, bg=self.surface)
        utils_form_left.pack(fill="x", pady=(0, 20))

        self._section_heading(utils_form_left, "复制站模板处理",
                              "批量处理 Shopify 主题模板文件").pack(fill="x", pady=(0, 16))

        # 产品文件路径
        utils_csv_wrap = tk.Frame(utils_form_left, bg=self.surface)
        utils_csv_wrap.pack(fill="x", pady=(0, 16))

        tk.Label(utils_csv_wrap, text="产品文件路径",
                 bg=self.surface, fg=self.text,
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w")

        utils_csv_row = tk.Frame(utils_csv_wrap, bg=self.surface)
        utils_csv_row.pack(fill="x", pady=(6, 0))

        self.utils_csv_var = tk.StringVar()
        utils_csv_entry = tk.Entry(utils_csv_row, textvariable=self.utils_csv_var,
                                    bg=self.field_bg, fg=self.text,
                                    relief="flat", insertwidth=2,
                                    insertbackground=self.brand_color,
                                    highlightthickness=1,
                                    highlightbackground=self.field_border,
                                    highlightcolor=self.brand_color,
                                    font=("Microsoft YaHei UI", 10))
        utils_csv_entry.pack(side="left", fill="x", expand=True, ipady=8, ipadx=10)

        utils_csv_btn = tk.Button(utils_csv_row, text="浏览",
                                   command=self._browse_utils_csv,
                                   bg=self.surface_alt, fg=self.text,
                                   activebackground=self.line,
                                   activeforeground=self.text,
                                   relief="flat", font=("Microsoft YaHei UI", 9, "bold"),
                                   cursor="hand2", padx=16, pady=8, bd=0,
                                   highlightthickness=0)
        utils_csv_btn.pack(side="left", padx=(10, 0))
        self._bind_hover(utils_csv_btn, self.surface_alt, self.line)

        # 主题压缩包
        utils_zip_wrap = tk.Frame(utils_form_left, bg=self.surface)
        utils_zip_wrap.pack(fill="x", pady=(0, 16))

        tk.Label(utils_zip_wrap, text="主题压缩包 (.zip)",
                 bg=self.surface, fg=self.text,
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w")

        utils_zip_row = tk.Frame(utils_zip_wrap, bg=self.surface)
        utils_zip_row.pack(fill="x", pady=(6, 0))

        self.utils_zip_var = tk.StringVar()
        utils_zip_entry = tk.Entry(utils_zip_row, textvariable=self.utils_zip_var,
                                    bg=self.field_bg, fg=self.text,
                                    relief="flat", insertwidth=2,
                                    insertbackground=self.brand_color,
                                    highlightthickness=1,
                                    highlightbackground=self.field_border,
                                    highlightcolor=self.brand_color,
                                    font=("Microsoft YaHei UI", 10))
        utils_zip_entry.pack(side="left", fill="x", expand=True, ipady=8, ipadx=10)

        utils_zip_btn = tk.Button(utils_zip_row, text="浏览",
                                   command=self._browse_utils_zip,
                                   bg=self.surface_alt, fg=self.text,
                                   activebackground=self.line,
                                   activeforeground=self.text,
                                   relief="flat", font=("Microsoft YaHei UI", 9, "bold"),
                                   cursor="hand2", padx=16, pady=8, bd=0,
                                   highlightthickness=0)
        utils_zip_btn.pack(side="left", padx=(10, 0))
        self._bind_hover(utils_zip_btn, self.surface_alt, self.line)

        # 新域名
        utils_domain_wrap = tk.Frame(utils_form_left, bg=self.surface)
        utils_domain_wrap.pack(fill="x", pady=(0, 16))

        tk.Label(utils_domain_wrap, text="新域名",
                 bg=self.surface, fg=self.text,
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w")

        self.utils_domain_var = tk.StringVar()
        utils_domain_entry = tk.Entry(utils_domain_wrap, textvariable=self.utils_domain_var,
                                       bg=self.field_bg, fg=self.text,
                                       relief="flat", insertwidth=2,
                                       insertbackground=self.brand_color,
                                       highlightthickness=1,
                                       highlightbackground=self.field_border,
                                       highlightcolor=self.brand_color,
                                       font=("Microsoft YaHei UI", 10))
        utils_domain_entry.pack(fill="x", pady=(6, 0), ipady=8, ipadx=10)
        tk.Label(utils_domain_wrap, text="例如：example.com",
                 bg=self.surface, fg=self.text_soft,
                 font=("Microsoft YaHei UI", 8)).pack(anchor="w", pady=(4, 0))

        # 开始处理按钮
        self.utils_replace_btn = tk.Button(utils_form_left, text="开始处理",
                                            command=self._on_utils_process,
                                            bg=self.brand_color, fg="#FFFFFF",
                                            activebackground=self.brand_press,
                                            activeforeground="#FFFFFF",
                                            relief="flat", font=("Microsoft YaHei UI", 10, "bold"),
                                            cursor="hand2", padx=24, pady=12, bd=0,
                                            highlightthickness=0)
        self.utils_replace_btn.pack(fill="x", ipady=4, pady=(20, 8))
        self._bind_hover(self.utils_replace_btn, self.brand_color, self.brand_hover)

        # 次要操作
        utils_secondary = tk.Frame(utils_form_left, bg=self.surface)
        utils_secondary.pack(fill="x")

        self.utils_open_file_btn = tk.Button(utils_secondary, text="打开文件",
                                              command=self._open_utils_output,
                                              bg=self.surface, fg=self.text_mid,
                                              activebackground=self.surface_alt,
                                              activeforeground=self.text,
                                              relief="flat", font=("Microsoft YaHei UI", 9),
                                              cursor="hand2", bd=0,
                                              highlightthickness=0)
        self.utils_open_file_btn.pack(side="left")
        self._bind_text_hover(self.utils_open_file_btn, self.text_mid, self.brand_color)

        tk.Label(utils_secondary, text="·", bg=self.surface, fg=self.text_soft,
                 font=("Microsoft YaHei UI", 9)).pack(side="left", padx=10)

        self.utils_open_folder_btn = tk.Button(utils_secondary, text="打开文件夹",
                                                command=self._open_utils_folder,
                                                bg=self.surface, fg=self.text_mid,
                                                activebackground=self.surface_alt,
                                                activeforeground=self.text,
                                                relief="flat", font=("Microsoft YaHei UI", 9),
                                                cursor="hand2", bd=0,
                                                highlightthickness=0)
        self.utils_open_folder_btn.pack(side="left")
        self._bind_text_hover(self.utils_open_folder_btn, self.text_mid, self.brand_color)

        # 右列
        utils_form_right = tk.Frame(utils_right_col, bg=self.surface)
        utils_form_right.pack(fill="x", pady=(0, 20))

        # 占位，对齐顶部
        tk.Frame(utils_form_right, bg=self.surface, height=50).pack(fill="x")

        # 替换 utils.js 选项
        utils_js_wrap = tk.LabelFrame(utils_form_right, text="替换 utils.js",
                                       bg=self.surface, fg=self.text,
                                       font=("Microsoft YaHei UI", 10, "bold"),
                                       padx=14, pady=10, bd=1, relief="solid")
        utils_js_wrap.pack(fill="x", pady=(0, 16))

        self.utils_replace_js_var = tk.BooleanVar(value=True)
        utils_js_check = tk.Checkbutton(utils_js_wrap, text="启用",
                                         variable=self.utils_replace_js_var,
                                         bg=self.surface, fg=self.text,
                                         activebackground=self.surface,
                                         activeforeground=self.text,
                                         selectcolor=self.field_bg,
                                         font=("Microsoft YaHei UI", 10),
                                         bd=0, highlightthickness=0,
                                         cursor="hand2")
        utils_js_check.pack(anchor="w")

        utils_js_file_row = tk.Frame(utils_js_wrap, bg=self.surface)
        utils_js_file_row.pack(fill="x", pady=(8, 0))

        tk.Label(utils_js_file_row, text="utils.js 文件：",
                 bg=self.surface, fg=self.text_mid,
                 font=("Microsoft YaHei UI", 9)).pack(side="left")

        self.utils_js_path_var = tk.StringVar()
        default_utils_js = os.path.join(RESOURCE_PATH, "data", "utils.js")
        if os.path.exists(default_utils_js):
            self.utils_js_path_var.set(default_utils_js)
        utils_js_entry = tk.Entry(utils_js_file_row, textvariable=self.utils_js_path_var,
                                   bg=self.field_bg, fg=self.text,
                                   relief="flat", insertwidth=2,
                                   insertbackground=self.brand_color,
                                   highlightthickness=1,
                                   highlightbackground=self.field_border,
                                   highlightcolor=self.brand_color,
                                   font=("Microsoft YaHei UI", 9))
        utils_js_entry.pack(side="left", fill="x", expand=True, padx=(6, 0), ipady=4, ipadx=6)

        utils_js_browse_btn = tk.Button(utils_js_file_row, text="浏览",
                                         command=self._browse_utils_js,
                                         bg=self.surface_alt, fg=self.text,
                                         activebackground=self.line,
                                         activeforeground=self.text,
                                         relief="flat", font=("Microsoft YaHei UI", 8, "bold"),
                                         cursor="hand2", padx=12, pady=4, bd=0,
                                         highlightthickness=0)
        utils_js_browse_btn.pack(side="left", padx=(8, 0))
        self._bind_hover(utils_js_browse_btn, self.surface_alt, self.line)

        utils_js_target_row = tk.Frame(utils_js_wrap, bg=self.surface)
        utils_js_target_row.pack(fill="x", pady=(6, 0))

        tk.Label(utils_js_target_row, text="目标路径：",
                 bg=self.surface, fg=self.text_mid,
                 font=("Microsoft YaHei UI", 9)).pack(side="left")

        self.utils_js_target_var = tk.StringVar(value="assets/utils.js")
        utils_js_target_entry = tk.Entry(utils_js_target_row, textvariable=self.utils_js_target_var,
                                          bg=self.field_bg, fg=self.text,
                                          relief="flat", insertwidth=2,
                                          insertbackground=self.brand_color,
                                          highlightthickness=1,
                                          highlightbackground=self.field_border,
                                          highlightcolor=self.brand_color,
                                          font=("Microsoft YaHei UI", 9))
        utils_js_target_entry.pack(side="left", fill="x", expand=True, padx=(6, 0), ipady=4, ipadx=6)

        # 处理选项说明
        utils_options_wrap = tk.LabelFrame(utils_form_right, text="将执行以下操作",
                                           bg=self.surface, fg=self.text,
                                           font=("Microsoft YaHei UI", 10, "bold"),
                                           padx=14, pady=10, bd=1, relief="solid")
        utils_options_wrap.pack(fill="x", pady=(0, 16))

        option_items = [
            "product.json — 添加自定义 Tabs + 随机销量数",
            "settings_data.json — 快速预览销量/浏览数",
            "footer-group.json — 版权域名替换",
            "assets/utils.js — 替换 utils.js 文件（可选）",
        ]
        for item in option_items:
            opt_row = tk.Frame(utils_options_wrap, bg=self.surface)
            opt_row.pack(fill="x", pady=2)
            tk.Label(opt_row, text="✓", bg=self.surface, fg=self.success_color,
                     font=("Microsoft YaHei UI", 9, "bold")).pack(side="left")
            tk.Label(opt_row, text="  " + item, bg=self.surface, fg=self.text_mid,
                     font=("Microsoft YaHei UI", 9)).pack(side="left")

        # 隐藏的输出文本框（保留代码引用，不显示）
        self.utils_output_text = tk.Text(self.utils_workspace, bg=self.output_bg, fg=self.output_text,
                                          relief="flat", bd=0, padx=16, pady=14,
                                          font=("Consolas", 10),
                                          wrap="word", cursor="arrow",
                                          state="disabled")
        self.utils_output_text.tag_configure("ok", foreground=self.success_color,
                                              font=("Consolas", 10, "bold"))
        self.utils_output_text.tag_configure("err", foreground="#E5484D",
                                              font=("Consolas", 10, "bold"))
        self.utils_output_text.tag_configure("warn", foreground="#D97706",
                                              font=("Consolas", 10, "bold"))
        self.utils_output_text.tag_configure("muted", foreground=self.text_soft)
        self.utils_output_text.tag_configure("accent", foreground="#7CC4FF",
                                              font=("Consolas", 10, "bold"))
        self.utils_output_text.tag_configure("line", foreground="#2A3038")

        self._write_utils_initial()

        self._last_utils_output = None

        # ===== TERMS WORKSPACE =====
        self.terms_workspace = tk.Frame(app, bg=self.surface)

        terms_cols = tk.Frame(self.terms_workspace, bg=self.surface)
        terms_cols.pack(fill="both", expand=True)

        # 左列：输入区
        terms_left = tk.Frame(terms_cols, bg=self.surface, width=460)
        terms_left.pack(side="left", fill="y", padx=(0, 24))
        terms_left.pack_propagate(False)

        terms_form = tk.Frame(terms_left, bg=self.surface)
        terms_form.pack(fill="x", pady=(0, 20))

        self._section_heading(terms_form, "条款页面生成",
                              "根据品牌名生成条款页面（带域名版本）").pack(fill="x", pady=(0, 16))

        # 品牌名
        self.terms_brand_var = tk.StringVar()
        brand_wrap, self.terms_brand_entry = self._labelled_field(terms_form, "品牌名", self.terms_brand_var, hint="例如：BAGSMART")
        brand_wrap.pack(fill="x", pady=(0, 16))

        # 域名
        self.terms_domain_var = tk.StringVar()
        domain_wrap, self.terms_domain_entry = self._labelled_field(terms_form, "域名", self.terms_domain_var, hint="例如：example.com")
        domain_wrap.pack(fill="x", pady=(0, 16))

        # 生成按钮
        self.terms_generate_btn = tk.Button(terms_form, text="开始生成",
                                            command=self._on_generate_terms,
                                            bg=self.brand_color, fg="#FFFFFF",
                                            activebackground=self.brand_press,
                                            activeforeground="#FFFFFF",
                                            relief="flat", font=("Microsoft YaHei UI", 10, "bold"),
                                            cursor="hand2", padx=24, pady=12, bd=0,
                                            highlightthickness=0)
        self.terms_generate_btn.pack(fill="x", ipady=4, pady=(8, 8))
        self._bind_hover(self.terms_generate_btn, self.brand_color, self.brand_hover)

        # 次要操作
        terms_secondary = tk.Frame(terms_form, bg=self.surface)
        terms_secondary.pack(fill="x")

        terms_open_btn = tk.Button(terms_secondary, text="打开文件",
                                    command=self._open_terms_output,
                                    bg=self.surface, fg=self.text_mid,
                                    activebackground=self.surface, activeforeground=self.brand_color,
                                    relief="flat", font=("Microsoft YaHei UI", 9),
                                    cursor="hand2", bd=0, highlightthickness=0)
        terms_open_btn.pack(side="left")
        self._bind_text_hover(terms_open_btn, self.text_mid, self.brand_color)

        tk.Label(terms_secondary, text="\u00b7", bg=self.surface, fg=self.text_soft,
                 font=("Microsoft YaHei UI", 9)).pack(side="left", padx=10)

        terms_folder_btn = tk.Button(terms_secondary, text="打开文件夹",
                                      command=self._open_terms_folder,
                                      bg=self.surface, fg=self.text_mid,
                                      activebackground=self.surface, activeforeground=self.brand_color,
                                      relief="flat", font=("Microsoft YaHei UI", 9),
                                      cursor="hand2", bd=0, highlightthickness=0)
        terms_folder_btn.pack(side="left")
        self._bind_text_hover(terms_folder_btn, self.text_mid, self.brand_color)

        # 右列：输出区
        terms_right = tk.Frame(terms_cols, bg=self.surface)
        terms_right.pack(side="left", fill="both", expand=True)

        tk.Label(terms_right, text="输出", bg=self.surface, fg=self.text,
                 font=("Microsoft YaHei UI", 11, "bold")).pack(anchor="w")

        out_container = tk.Frame(terms_right, bg=self.output_bg, highlightthickness=0)
        out_container.pack(fill="both", expand=True, pady=(8, 0))

        self.terms_output_text = tk.Text(out_container, wrap="word",
                                          bg=self.output_bg, fg=self.output_text,
                                          relief="flat", font=("Consolas", 11),
                                          padx=20, pady=16,
                                          insertbackground=self.output_text,
                                          highlightthickness=0,
                                          selectbackground="#2A3441",
                                          selectforeground="#FFFFFF",
                                          cursor="xterm")
        self.terms_output_text.pack(fill="both", expand=True)

        self.terms_output_text.tag_configure("muted", foreground="#6D7B8A", font=("Consolas", 11))
        self.terms_output_text.tag_configure("accent", foreground="#4CB3FF", font=("Consolas", 11, "bold"))
        self.terms_output_text.tag_configure("ok", foreground="#5FD7A2", font=("Consolas", 11, "bold"))
        self.terms_output_text.tag_configure("warn", foreground="#FFB454", font=("Consolas", 11, "bold"))
        self.terms_output_text.tag_configure("err", foreground="#FF7A7A", font=("Consolas", 11, "bold"))
        self.terms_output_text.tag_configure("line", foreground="#3A4653", font=("Consolas", 11))
        self.terms_output_text.tag_configure("key", foreground="#D4A5FF", font=("Consolas", 11))

        self._write_terms_initial()

        self._last_terms_output = None

        # STATUS BAR
        sb = tk.Frame(root, bg="#FFFFFF", height=24,
                       highlightbackground=self.line, highlightthickness=1)
        sb.pack(fill="x", side="bottom")
        sb.pack_propagate(False)
        tk.Label(sb, text="  豆脚Altera Excel 生成工具",
                 bg="#FFFFFF", fg=self.text_soft,
                 font=("Microsoft YaHei UI", 8),
                 anchor="w").pack(fill="x", expand=True, side="left")

        self._last_output_path = None

    # ---- UI helpers ----
    def _section_heading(self, parent, title, subtitle):
        wrap = tk.Frame(parent, bg=self.surface)
        tk.Label(wrap, text=title, bg=self.surface, fg=self.text,
                 font=("Microsoft YaHei UI", 11, "bold")).pack(anchor="w")
        tk.Label(wrap, text=subtitle, bg=self.surface, fg=self.text_soft,
                 font=("Microsoft YaHei UI", 9)).pack(anchor="w", pady=(2, 0))
        return wrap

    def _labelled_field(self, parent, label, var, hint=""):
        wrap = tk.Frame(parent, bg=self.surface)
        tk.Label(wrap, text=label, bg=self.surface, fg=self.text,
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w")
        entry = tk.Entry(wrap, textvariable=var,
                          bg=self.field_bg, fg=self.text,
                          relief="flat", insertwidth=2,
                          insertbackground=self.brand_color,
                          highlightthickness=1,
                          highlightbackground=self.field_border,
                          highlightcolor=self.brand_color,
                          font=("Microsoft YaHei UI", 10))
        entry.pack(fill="x", pady=(6, 0), ipady=8, ipadx=10)
        if hint:
            tk.Label(wrap, text=hint, bg=self.surface, fg=self.text_soft,
                     font=("Microsoft YaHei UI", 8)).pack(anchor="w", pady=(4, 0))
        return wrap, entry

    def _bind_hover(self, widget, bg_rest, bg_hover):
        widget.bind("<Enter>", lambda e: widget.configure(bg=bg_hover))
        widget.bind("<Leave>", lambda e: widget.configure(bg=bg_rest))

    def _bind_text_hover(self, widget, fg_rest, fg_hover):
        widget.bind("<Enter>", lambda e: widget.configure(fg=fg_hover))
        widget.bind("<Leave>", lambda e: widget.configure(fg=fg_rest))

    def _write_initial(self):
        segs = [
            ("\u2713 就绪\n\n", "ok"),
            ("请在左侧填写：\n", "muted"),
            ("  1. 产品文件路径（可点击\"浏览\"选择）\n", "muted"),
            ("  2. 品牌、标题、描述、域名\n", "muted"),
            ("  3. 点击 \"开始处理\"\n\n", "muted"),
            ("\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n\n", "line"),
            ("将生成 4 个工作表：\n", "muted"),
            ("  \u2022 ", "key"), ("Metafield Definitions", "accent"), ("   从 CSV 列头生成\n", "muted"),
            ("  \u2022 ", "key"), ("Smart Collections    ", "accent"), ("   从 Tags 列生成\n", "muted"),
            ("  \u2022 ", "key"), ("Pages                ", "accent"), ("   根据品牌信息生成\n", "muted"),
            ("  \u2022 ", "key"), ("Menu                 ", "accent"), ("   分类菜单 + Policy/About 组\n", "muted"),
        ]
        for text, tag in segs:
            if tag:
                self.result_text.insert("end", text, tag)
            else:
                self.result_text.insert("end", text)

    def _append_result_lines(self, tuples_list):
        self.result_text.configure(state="normal")
        for text, tag in tuples_list:
            if tag:
                self.result_text.insert("end", text, tag)
            else:
                self.result_text.insert("end", text)
        self.result_text.see("end")
        self.result_text.configure(state="disabled")

    def _format_menu_tree(self, menu_rows):
        """将菜单按 Handle 分组后，按 Parent ID 层级生成树形输出"""
        lines = []
        lines.append(("\u2500" * 50 + "\n", "line"))
        lines.append(("  \u83dc\u5355\u7ed3\u6784\n", "key"))
        lines.append(("\u2500" * 50 + "\n", "line"))

        # 按 Handle 分组：main-menu / policy / about
        by_handle = {}
        for row in menu_rows:
            h = row.get("Handle", "")
            if h not in by_handle:
                by_handle[h] = []
            by_handle[h].append(row)

        menu_labels = {
            "main-menu": "Main Menu (\u5546\u54c1\u5206\u7c7b)",
            "policy": "Policy (\u6761\u6b3e)",
            "about": "About (\u5173\u4e8e)",
        }

        for handle in ["main-menu", "policy", "about"]:
            if handle not in by_handle:
                continue
            items = by_handle[handle]
            lines.append(("\n", None))
            lines.append((f"  \u25a1 {menu_labels.get(handle, handle)}  [{len(items)}]\n", "accent"))

            # 构建 id -> row 的索引
            id_to_row = {row["Menu Item: ID"]: row for row in items}
            # 构建 parent -> children 列表
            children_of = {}
            for row in items:
                p_id = row["Menu Item: Parent ID"]
                if p_id not in children_of:
                    children_of[p_id] = []
                children_of[p_id].append(row)

            # 计算每个节点的层级（深度）
            def _depth(item_id, memo):
                if item_id in memo:
                    return memo[item_id]
                row = id_to_row.get(item_id)
                if not row:
                    memo[item_id] = 0
                    return 0
                p_id = row["Menu Item: Parent ID"]
                if not p_id or p_id not in id_to_row:
                    memo[item_id] = 1
                    return 1
                d = 1 + _depth(p_id, memo)
                memo[item_id] = d
                return d

            depth_memo = {}
            for row in items:
                _depth(row["Menu Item: ID"], depth_memo)

            # 按深度递归打印
            def _print_children(parent_id, prefix, depth):
                kids = children_of.get(parent_id, [])
                for i, kid in enumerate(kids):
                    is_last = (i == len(kids) - 1)
                    connector = "\u2514\u2500 " if is_last else "\u251c\u2500 "
                    title = kid["Menu Item: Title"]
                    url = kid.get("Menu Item: URL", "")
                    # 深度 1 的节点用特殊符号
                    if depth == 1:
                        # 一级分类：用更粗的符号
                        lines.append((f"  {prefix}{connector}{title}\n", "key"))
                    elif depth == 2:
                        lines.append((f"  {prefix}{connector}{title}\n", "muted"))
                    else:
                        lines.append((f"  {prefix}{connector}{title}\n", "muted"))

                    # 为子节点准备前缀
                    child_prefix = prefix + ("    " if is_last else "\u2502   ")
                    _print_children(kid["Menu Item: ID"], child_prefix, depth + 1)

            # 从根节点（Parent ID 为空）开始递归打印
            roots = children_of.get("", [])
            roots_sorted = sorted(roots, key=lambda r: int(r["Menu Item: ID"]))
            for i, root in enumerate(roots_sorted):
                is_last_root = (i == len(roots_sorted) - 1)
                root_connector = "\u2514\u2500 " if is_last_root else "\u251c\u2500 "
                root_title = root["Menu Item: Title"]
                root_has_children = root["Menu Item: ID"] in children_of and len(children_of[root["Menu Item: ID"]]) > 0
                lines.append((f"  {root_connector}{root_title}\n", "key"))
                if root_has_children:
                    child_prefix = "    " if is_last_root else "\u2502   "
                    _print_children(root["Menu Item: ID"], child_prefix, 2)

        lines.append(("\n", None))
        return lines

    # ---- Business helpers ----
    def _auto_find_shopify_csv(self):
        script_dir = RESOURCE_PATH
        pattern = re.compile(r".*_shopify_.*\.csv$", re.IGNORECASE)
        for f in os.listdir(script_dir):
            full_path = os.path.join(script_dir, f)
            if os.path.isfile(full_path) and pattern.match(f):
                return full_path
        return None

    def _browse_csv(self):
        path = filedialog.askopenfilename(
            title="选择产品 CSV 文件",
            filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")]
        )
        if path:
            self.csv_path_var.set(path)
            # 自动识别品牌名：取文件名（不含扩展名）按 _ 分割的第一个部分
            import os
            filename = os.path.splitext(os.path.basename(path))[0]
            parts = filename.split("_")
            if parts and parts[0] and not parts[0].isdigit():
                self.brand_var.set(parts[0].capitalize())

    def _browse_utils_csv(self):
        path = filedialog.askopenfilename(
            title="选择产品文件",
            filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")]
        )
        if path:
            self.utils_csv_var.set(path)
            self._utils_append_output(f"已选择产品文件: {os.path.basename(path)}\n", "muted")

    def _browse_utils_zip(self):
        path = filedialog.askopenfilename(
            title="选择主题压缩包",
            filetypes=[("ZIP 压缩包", "*.zip"), ("所有文件", "*.*")]
        )
        if path:
            self.utils_zip_var.set(path)
            self._utils_append_output(f"已选择主题包: {os.path.basename(path)}\n", "muted")

    def _browse_utils_js(self):
        path = filedialog.askopenfilename(
            title="选择 utils.js 文件",
            filetypes=[("JS 文件", "*.js"), ("所有文件", "*.*")]
        )
        if path:
            self.utils_js_path_var.set(path)
            self._utils_append_output(f"已选择 utils.js: {os.path.basename(path)}\n", "muted")

    def _write_utils_initial(self):
        self._utils_append_output("\u2713 就绪\n\n", "ok")
        self._utils_append_output("请在左侧填写：\n", "muted")
        self._utils_append_output("  1. 产品文件路径（可点击\"浏览\"选择）\n", "muted")
        self._utils_append_output("  2. 主题压缩包 (.zip)\n", "muted")
        self._utils_append_output("  3. 新域名\n", "muted")
        self._utils_append_output("  4. 点击 \"开始处理\"\n\n", "muted")
        self._utils_append_output("─" * 35 + "\n\n", "line")
        self._utils_append_output("将处理 3 个模板文件：\n", "muted")
        self._utils_append_output("  \u2022 ", "key")
        self._utils_append_output("product.json", "accent")
        self._utils_append_output("   自定义 Tabs + 随机销量\n", "muted")
        self._utils_append_output("  \u2022 ", "key")
        self._utils_append_output("settings_data.json", "accent")
        self._utils_append_output("   快速预览销量/浏览数\n", "muted")
        self._utils_append_output("  \u2022 ", "key")
        self._utils_append_output("footer-group.json", "accent")
        self._utils_append_output("   版权域名替换\n", "muted")
        self._utils_append_output("  \u2022 ", "key")
        self._utils_append_output("assets/utils.js", "accent")
        self._utils_append_output("   替换 utils.js（可选）\n", "muted")

    def _utils_append_output(self, text, tag="muted"):
        self.utils_output_text.configure(state="normal")
        self.utils_output_text.insert("end", text, tag)
        self.utils_output_text.see("end")
        self.utils_output_text.configure(state="disabled")

    def _utils_clear_output(self):
        self.utils_output_text.configure(state="normal")
        self.utils_output_text.delete("1.0", "end")
        self.utils_output_text.configure(state="disabled")

    def _on_utils_process(self):
        csv_path = self.utils_csv_var.get().strip()
        zip_path = self.utils_zip_var.get().strip()
        domain = self.utils_domain_var.get().strip()
        replace_js = self.utils_replace_js_var.get()
        js_file = self.utils_js_path_var.get().strip()
        js_target = self.utils_js_target_var.get().strip()

        if not csv_path:
            messagebox.showwarning("提示", "请选择产品文件")
            return
        if not os.path.exists(csv_path):
            messagebox.showerror("错误", "产品文件不存在")
            return
        if not zip_path:
            messagebox.showwarning("提示", "请选择主题压缩包")
            return
        if not os.path.exists(zip_path):
            messagebox.showerror("错误", "主题压缩包不存在")
            return
        if not domain:
            messagebox.showwarning("提示", "请填写新域名")
            return
        if replace_js:
            if not js_file:
                messagebox.showwarning("提示", "请选择 utils.js 文件")
                return
            if not os.path.exists(js_file):
                messagebox.showerror("错误", "utils.js 文件不存在")
                return
            if not js_target:
                messagebox.showwarning("提示", "请填写 utils.js 目标路径")
                return

        self._last_utils_output = zip_path
        self.utils_replace_btn.configure(state="disabled", text="处理中...")
        self._utils_clear_output()
        self._utils_append_output("开始处理...\n\n", "ok")

        try:
            self._do_utils_process(csv_path, zip_path, domain, replace_js, js_file, js_target)
            self._utils_append_output("\n\u2713 全部完成！\n", "ok")
            messagebox.showinfo("成功", "模板处理完成！\n压缩包已原地更新。")
        except Exception as e:
            self._utils_append_output(f"\n\u2717 处理失败: {str(e)}\n", "err")
            messagebox.showerror("错误", f"处理失败：{str(e)}")
        finally:
            self.utils_replace_btn.configure(state="normal", text="开始处理")

    def _do_utils_process(self, csv_path, zip_path, domain, replace_js=False, js_file=None, js_target=None):
        temp_dir = tempfile.mkdtemp()
        try:
            total_steps = 4 + (1 if replace_js else 0)
            self._utils_append_output(f"[1/{total_steps}] 解压主题包...\n", "muted")
            extract_dir = os.path.join(temp_dir, "theme")
            os.makedirs(extract_dir)
            with zipfile.ZipFile(zip_path, 'r') as zf:
                zf.extractall(extract_dir)
            self._utils_append_output("      解压完成\n\n", "muted")

            root_dir = self._find_theme_root(extract_dir)
            product_json = os.path.join(root_dir, "templates", "product.json")
            settings_json = os.path.join(root_dir, "config", "settings_data.json")
            footer_json = os.path.join(root_dir, "sections", "footer-group.json")

            self._utils_append_output(f"[2/{total_steps}] 读取产品数据...\n", "muted")
            metafield_rows = build_metafield_definitions_from_csv(csv_path)
            self._utils_append_output(f"      自定义列: {len(metafield_rows)} 个\n\n", "muted")

            self._utils_append_output(f"[3/{total_steps}] 处理 product.json...\n", "muted")
            if os.path.exists(product_json):
                build_custom_tabs_in_product_json(product_json, metafield_rows)
                self._utils_append_output("      ✓ 已添加自定义 Tabs\n", "muted")
                randomize_sold_in_number(product_json)
                self._utils_append_output("      ✓ 已生成随机销量数\n\n", "muted")
            else:
                self._utils_append_output("      ⚠ 未找到 product.json，跳过\n\n", "warn")

            step = 4
            self._utils_append_output(f"[{step}/{total_steps}] 处理 settings_data.json + footer...\n", "muted")
            if os.path.exists(settings_json):
                randomize_settings_data(settings_json, hot_keywords=None)
                self._utils_append_output("      ✓ 已生成快速预览销量/浏览数\n", "muted")
            else:
                self._utils_append_output("      ⚠ 未找到 settings_data.json，跳过\n", "warn")

            if os.path.exists(footer_json):
                replace_footer_copyright_domain(footer_json, domain)
                self._utils_append_output("      ✓ 已替换版权域名\n\n", "muted")
            else:
                self._utils_append_output("      ⚠ 未找到 footer-group.json，跳过\n\n", "warn")

            if replace_js:
                step += 1
                self._utils_append_output(f"[{step}/{total_steps}] 替换 utils.js...\n", "muted")
                target_full = os.path.join(root_dir, js_target)
                target_dir = os.path.dirname(target_full)
                if not os.path.exists(target_dir):
                    os.makedirs(target_dir, exist_ok=True)
                import shutil as _shutil
                _shutil.copy2(js_file, target_full)
                self._utils_append_output(f"      ✓ 已替换 {js_target}\n\n", "muted")

            self._utils_append_output("重新打包压缩包...\n", "muted")
            temp_zip = os.path.join(temp_dir, "output.zip")
            with zipfile.ZipFile(temp_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                for root, dirs, files in os.walk(root_dir):
                    for f in files:
                        full = os.path.join(root, f)
                        arcname = os.path.relpath(full, root_dir)
                        zf.write(full, arcname)
            shutil.move(temp_zip, zip_path)
            self._utils_append_output("      ✓ 打包完成\n", "muted")

        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def _find_theme_root(self, extract_dir):
        entries = os.listdir(extract_dir)
        if len(entries) == 1 and os.path.isdir(os.path.join(extract_dir, entries[0])):
            return os.path.join(extract_dir, entries[0])
        return extract_dir

    # ---- 条款页面生成 ----

    def _write_terms_initial(self):
        SEP = "\u2500" * 45
        self._append_terms_lines([
            ("\u2713 就绪\n\n", "ok"),
            ("请在左侧填写品牌名和域名，点击\"开始生成\"\n", "muted"),
            ("将生成带域名的条款页面（可点击链接）\n\n", "muted"),
            (SEP + "\n\n", "line"),
            ("将生成 8 个条款页面：\n", "muted"),
            ("  \u2022 ", "key"), ("shipping-policy       ", "accent"), ("运输政策\n", "muted"),
            ("  \u2022 ", "key"), ("refund-policy         ", "accent"), ("退款政策\n", "muted"),
            ("  \u2022 ", "key"), ("privacy-policy        ", "accent"), ("隐私政策\n", "muted"),
            ("  \u2022 ", "key"), ("customer-service-policy", "accent"), ("客服政策\n", "muted"),
            ("  \u2022 ", "key"), ("terms-of-purchase     ", "accent"), ("购买条款\n", "muted"),
            ("  \u2022 ", "key"), ("terms-of-use          ", "accent"), ("使用条款\n", "muted"),
            ("  \u2022 ", "key"), ("about-us              ", "accent"), ("关于我们\n", "muted"),
            ("  \u2022 ", "key"), ("contact-us            ", "accent"), ("联系我们\n", "muted"),
        ])

    def _append_terms_lines(self, tuples_list):
        self.terms_output_text.configure(state="normal")
        for text, tag in tuples_list:
            self.terms_output_text.insert("end", text, tag)
        self.terms_output_text.see("end")
        self.terms_output_text.configure(state="disabled")

    def _on_generate_terms(self):
        brand = self.terms_brand_var.get().strip()
        domain = self.terms_domain_var.get().strip()

        if not brand:
            messagebox.showwarning("提示", "请填写品牌名")
            return
        if not domain:
            messagebox.showwarning("提示", "请填写域名")
            return

        self.terms_output_text.configure(state="normal")
        self.terms_output_text.delete("1.0", "end")
        self.terms_output_text.configure(state="disabled")

        self._append_terms_lines([
            ("\u27f3 正在生成条款页面...\n\n", "warn"),
        ])

        try:
            domain_clean = domain.replace("https://", "").replace("http://", "").split("/")[0]
            domain_link_html = f'<a href="https://{domain_clean}" target="_blank" style="color: inherit; text-decoration: underline;">{domain_clean}</a>'

            ctx = {
                "domain": domain_clean,
                "brand": brand,
                "domain_link": domain_link_html,
                "email_press": f"press@{domain_clean}",
                "email_support": f"support@{domain_clean}",
                "email_wholesale": f"wholesale@{domain_clean}",
            }

            pages = _get_pages(with_domain=True)
            page_order = [
                "shipping-policy", "refund-policy", "privacy-policy",
                "customer-service-policy", "terms-of-purchase", "terms-of-use",
                "about-us", "contact-us",
            ]

            # 创建输出目录
            safe_brand = "".join(c for c in brand if c.isalnum() or c in "-_")[:40] or "store"
            output_dir = os.path.join(os.getcwd(), f"{safe_brand}_terms")
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, f"{safe_brand}_pages.xlsx")

            SEP = "\u2500" * 45

            self._append_terms_lines([
                (SEP + "\n", "line"),
                ("  版本选择明细\n", "key"),
                (SEP + "\n\n", "line"),
            ])

            pages_rows = []
            for handle in page_order:
                versions = pages.get(handle)
                if not versions:
                    self._append_terms_lines([
                        (f"  {handle:30s}", "muted"),
                        ("\u2717 无可用版本\n", "err"),
                    ])
                    continue

                page_def = random.choice(versions)
                source_file = page_def.get("_source_file", "unknown")
                page_title = page_def.get("title", handle.replace("-", " ").title())

                # 渲染页面 HTML
                html_content = _render_page(page_def, ctx)
                if not html_content:
                    html_content = f"<h1>{page_title}</h1>"

                pages_rows.append({
                    "Handle": handle,
                    "Title": page_title,
                    "Body (HTML)": html_content,
                    "Template Suffix": "",
                    "Published": "TRUE",
                })

                self._append_terms_lines([
                    (f"  {handle:30s}", "key"),
                    ("\u2192 ", "muted"),
                    (f"{source_file}\n", "accent"),
                ])

            # 生成 xlsx
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            write_dict_rows_to_sheet(wb, "Pages", PAGES_COLUMNS, pages_rows)
            wb.save(output_path)

            self._last_terms_output = output_path

            self._append_terms_lines([
                ("\n" + SEP + "\n", "line"),
                ("\u2713 条款页面生成完成\n\n", "ok"),
                (f"  品牌: {brand}\n", "muted"),
                (f"  域名: {domain_clean}\n", "muted"),
                ("  版本: 带域名（可点击链接）\n\n", "muted"),
                (f"  文件: {os.path.basename(output_path)}\n", "accent"),
                (f"  路径: {output_dir}\n", "muted"),
                (f"  共 {len(pages_rows)} 个条款页面\n", "muted"),
            ])

        except Exception as e:
            self._append_terms_lines([
                (f"\n\u2717 生成失败: {str(e)}\n", "err"),
            ])
            messagebox.showerror("错误", f"生成失败：{str(e)}")

    def _open_terms_output(self):
        if self._last_terms_output and os.path.exists(self._last_terms_output):
            os.startfile(self._last_terms_output)
        else:
            messagebox.showinfo("提示", "暂无输出文件，请先生成条款页面")

    def _open_terms_folder(self):
        if self._last_terms_output and os.path.exists(self._last_terms_output):
            os.startfile(os.path.dirname(self._last_terms_output))
        else:
            messagebox.showinfo("提示", "暂无输出文件，请先生成条款页面")

    def _open_utils_output(self):
        if self._last_utils_output and os.path.exists(self._last_utils_output):
            os.startfile(self._last_utils_output)
        else:
            messagebox.showinfo("提示", "暂无输出文件，请先处理模板")

    def _open_utils_folder(self):
        if self._last_utils_output and os.path.exists(self._last_utils_output):
            os.startfile(os.path.dirname(self._last_utils_output))
        else:
            messagebox.showinfo("提示", "暂无输出文件，请先处理模板")

    # ---- Event handlers ----
    def _parse_version(self, ver_str):
        try:
            return tuple(int(x) for x in ver_str.strip().split("."))
        except Exception:
            return (0, 0, 0)

    def _check_update_async(self):
        if not _HAS_URLLIB:
            self.root.after(0, lambda: messagebox.showinfo("检查更新", "当前环境不支持网络请求。"))
            return
        try:
            req = urlrequest.Request(UPDATE_CHECK_URL, headers={"User-Agent": "doujiao-updater"})
            with urlrequest.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            latest_ver = data.get("version", "0.0.0")
            download_url = data.get("green_download_url", "") or data.get("download_url", "")
            release_notes = data.get("release_notes", "")
            current = self._parse_version(APP_VERSION)
            latest = self._parse_version(latest_ver)
            has_update = latest > current
            self.root.after(0, lambda: self._show_update_result(has_update, latest_ver, download_url, release_notes))
        except Exception as e:
            self.root.after(0, lambda: messagebox.showwarning("检查更新", f"检查更新失败，请检查网络连接。\n\n错误信息：{e}"))

    def _show_update_result(self, has_update, latest_ver, download_url, release_notes):
        self.check_update_btn.config(state="normal", text="检查更新")
        if has_update:
            notes_text = f"\n\n{release_notes}" if release_notes else ""
            if messagebox.askyesno("发现新版本",
                                    f"当前版本：v{APP_VERSION}\n最新版本：v{latest_ver}{notes_text}\n\n是否立即下载更新？"):
                if download_url:
                    self._download_and_update(download_url)
        else:
            messagebox.showinfo("检查更新", f"当前已是最新版本！\n当前版本：v{APP_VERSION}")

    def _get_app_dir(self):
        if getattr(sys, 'frozen', False):
            return os.path.dirname(sys.executable)
        else:
            return os.path.dirname(os.path.abspath(__file__))

    def _download_and_update(self, download_url):
        import threading
        self.check_update_btn.config(state="disabled", text="下载中 0%")
        self._download_progress = 0

        def _do_download():
            import tempfile
            import zipfile
            tmp_dir = tempfile.mkdtemp(prefix="doujiao_update_")
            zip_path = os.path.join(tmp_dir, "update.zip")
            max_retries = 2
            last_error = None

            for attempt in range(max_retries + 1):
                try:
                    if os.path.exists(zip_path):
                        os.remove(zip_path)
                    req = urlrequest.Request(download_url, headers={"User-Agent": "doujiao-updater"})
                    with urlrequest.urlopen(req, timeout=120) as resp:
                        if resp.status != 200:
                            last_error = f"服务器返回错误：HTTP {resp.status}"
                            continue
                        total = int(resp.headers.get("Content-Length", 0))
                        downloaded = 0
                        with open(zip_path, "wb") as f:
                            while True:
                                chunk = resp.read(16384)
                                if not chunk:
                                    break
                                f.write(chunk)
                                downloaded += len(chunk)
                                if total > 0:
                                    pct = int(downloaded / total * 100)
                                    if pct != self._download_progress:
                                        self._download_progress = pct
                                        self.root.after(0, lambda p=pct: self.check_update_btn.config(text=f"下载中 {p}%"))
                        if total > 0 and downloaded != total:
                            last_error = f"下载不完整（{downloaded}/{total} 字节）"
                            continue
                        if not zipfile.is_zipfile(zip_path):
                            last_error = "更新包格式错误（文件可能已损坏）"
                            continue
                        with zipfile.ZipFile(zip_path, "r") as zf:
                            names = zf.namelist()
                            exe_name = "豆脚AlteraExcel工具.exe"
                            has_exe = any(n.endswith(exe_name) or n == exe_name for n in names)
                            if not has_exe:
                                last_error = "更新包缺少主程序文件"
                                continue
                    self.root.after(0, lambda: self._prepare_update(zip_path, tmp_dir))
                    return
                except Exception as e:
                    last_error = str(e)
                    if attempt < max_retries:
                        self.root.after(0, lambda: self.check_update_btn.config(text=f"重试中 ({attempt + 1}/{max_retries})"))
                        import time
                        time.sleep(1)

            try:
                import shutil
                if os.path.exists(zip_path) and os.path.getsize(zip_path) > 0:
                    try:
                        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
                        if not os.path.exists(desktop):
                            desktop = os.path.join(os.path.expanduser("~"), "桌面")
                        debug_file = os.path.join(desktop, "update_debug.zip")
                        shutil.copy2(zip_path, debug_file)
                    except Exception:
                        pass
                shutil.rmtree(tmp_dir, ignore_errors=True)
            except Exception:
                pass
            self.root.after(0, lambda: self._download_failed(last_error or "未知错误"))

        threading.Thread(target=_do_download, daemon=True).start()

    def _prepare_update(self, zip_path, tmp_dir):
        self.check_update_btn.config(state="normal", text="检查更新")
        try:
            import zipfile
            import subprocess
            extract_dir = os.path.join(tmp_dir, "extracted")
            os.makedirs(extract_dir, exist_ok=True)
            with zipfile.ZipFile(zip_path, "r") as zf:
                zf.extractall(extract_dir)
            exe_name = "豆脚AlteraExcel工具.exe"
            src_exe = os.path.join(extract_dir, exe_name)
            if not os.path.exists(src_exe):
                for root_dir, dirs, files in os.walk(extract_dir):
                    if exe_name in files:
                        src_exe = os.path.join(root_dir, exe_name)
                        extract_dir = root_dir
                        break
            if not os.path.exists(src_exe):
                messagebox.showerror("更新失败", "更新包格式不正确，缺少主程序文件")
                return
            app_dir = self._get_app_dir()
            if messagebox.askyesno("下载完成",
                                    f"更新包已下载完成，是否立即更新？\n\n更新后软件将自动重启。"):
                if getattr(sys, 'frozen', False):
                    current_exe = sys.executable
                else:
                    current_exe = sys.executable
                update_exe = os.path.join(tmp_dir, exe_name)
                shutil.copy2(current_exe, update_exe)
                subprocess.Popen([update_exe, "--do-update", extract_dir, app_dir, tmp_dir])
                self.root.after(300, self.root.destroy)
        except Exception as e:
            messagebox.showerror("更新失败", f"准备更新包失败：{e}")

    def _download_failed(self, err_msg):
        self.check_update_btn.config(state="normal", text="检查更新")
        messagebox.showerror("下载失败", f"下载更新包失败：{err_msg}")

    def _on_check_update(self):
        self.check_update_btn.config(state="disabled", text="检查中...")
        import threading
        threading.Thread(target=self._check_update_async, daemon=True).start()

    def on_generate(self):
        csv_path = self.csv_path_var.get().strip()
        brand = self.brand_var.get().strip()
        domain = self.domain_var.get().strip()
        page_version = self.page_version_var.get()
        gen_no_domain = page_version == "no_domain"
        gen_with_domain = page_version == "with_domain"

        self.result_text.configure(state="normal")
        self.result_text.delete("1.0", "end")
        self.result_text.configure(state="disabled")

        self._append_result_lines([
            ("⟳ 正在生成工作簿...\n\n", "warn"),
        ])
        self.root.update_idletasks()

        errors = []
        if not csv_path or not os.path.exists(csv_path):
            errors.append("CSV 文件不存在，请选择有效路径")
        if not brand: errors.append("品牌不能为空")
        if gen_with_domain and not domain:
            errors.append("选择带域名版本时，域名不能为空")

        if errors:
            self.result_text.configure(state="normal")
            self.result_text.delete("1.0", "end")
            self.result_text.configure(state="disabled")
            self._append_result_lines([("\u2717 输入错误\n\n", "err")])
            for e in errors:
                self._append_result_lines([("  \u25b8 ", "line"), (e + "\n", "muted")])
            messagebox.showerror("验证失败", "\n".join(errors))
            return

        try:
            results = []
            if gen_no_domain:
                result_no_domain = generate_excel(
                    domain=domain, brand=brand, product_csv_path=csv_path,
                    with_domain=False,
                )
                results.append(("不带域名", result_no_domain))

            if gen_with_domain:
                result_with_domain = generate_excel(
                    domain=domain, brand=brand, product_csv_path=csv_path,
                    with_domain=True,
                )
                results.append(("带域名", result_with_domain))

            # 使用最后一个主结果供后续操作（打开文件等）
            self._last_output_path = results[-1][1]["output_path"]

            self.result_text.configure(state="normal")
            self.result_text.delete("1.0", "end")
            self.result_text.configure(state="disabled")

            # 如果生成了两个版本，先输出概览
            if len(results) == 2:
                self._append_result_lines([
                    ("\u2713 生成成功（共 2 个版本）\n\n", "ok"),
                    ("\u2500" * 50 + "\n", "line"),
                    ("  版本输出路径：\n\n", "key"),
                ])
                for version_label, result in results:
                    self._append_result_lines([
                        (f"  \u26a0 {version_label}版本\n", "accent"),
                        (f"     文件: {os.path.basename(result['output_path'])}\n", "muted"),
                        (f"     路径: {os.path.dirname(result['output_path'])}\n\n", "muted"),
                    ])
                self._append_result_lines([
                    ("\u2500" * 50 + "\n", "line"),
                ])
            else:
                result = results[0][1]
                self._append_result_lines([
                    ("\u2713 生成成功\n\n", "ok"),
                    (os.path.basename(result["output_path"]) + "\n", "accent"),
                    (os.path.dirname(result["output_path"]) + "\n\n", "muted"),
                    ("\u2500" * 50 + "\n", "line"),
                ])

            # 输出第一个版本（或唯一版本）的详细信息
            primary_result = results[0][1]

            # ---- 文档行数统计 ----
            self._append_result_lines([
                ("  文档行数统计\n\n", "key"),
                ("  Metafield 定义   ", "key"), (f'{primary_result["metafield_count"]} 行\n', "muted"),
                ("  Smart Collections", "key"), (f'{primary_result["collection_count"]} 行\n', "muted"),
                ("  Pages            ", "key"), (f'{primary_result["pages_count"]} 行\n', "muted"),
                ("  Menu             ", "key"), (f'{primary_result["menu_count"]} 行\n', "muted"),
                ("\u2500" * 50 + "\n", "line"),
                ("\n  工作表拆分详情（Altera 单表上限 99 行，自动拆分）\n\n", "key"),
            ])

            # 输出每个 sheet 的拆分情况
            for base_name, total_rows, sheet_infos in primary_result["all_sheet_infos"]:
                if len(sheet_infos) > 1:
                    self._append_result_lines([
                        (f"  \u26a0 {base_name}", "accent"),
                        (f": 共 {total_rows} 行 → 拆分为 {len(sheet_infos)} 个 sheet\n", "muted"),
                    ])
                    for sheet_name, rows in sheet_infos:
                        self._append_result_lines([
                            (f"     \u2514\u2500 {sheet_name}", "key"),
                            (f": {rows} 行\n", "muted"),
                        ])
                else:
                    sheet_name, rows = sheet_infos[0]
                    self._append_result_lines([
                        (f"  \u2713 {base_name}", "ok"),
                        (f": {rows} 行 → 无需拆分\n", "muted"),
                    ])

            # 输出超出 99 行的 CSV 文件信息
            if primary_result.get("csv_files_info"):
                self._append_result_lines([
                    ("\n" + "\u2500" * 50 + "\n", "line"),
                    ("\n  超出 99 行部分已输出到 altera/ 文件夹（CSV 格式）\n\n", "key"),
                ])
                for base_name, files in primary_result["csv_files_info"]:
                    self._append_result_lines([
                        (f"  \u25cf {base_name}", "accent"),
                        (f": 共生成 {len(files)} 个 CSV 文件\n", "muted"),
                    ])
                    for fpath, frows in files:
                        self._append_result_lines([
                            (f"     \u2514\u2500 {os.path.basename(fpath)}", "key"),
                            (f": {frows} 行\n", "muted"),
                        ])
                altera_dir = os.path.join(os.path.dirname(primary_result["output_path"]), "altera")
                self._append_result_lines([
                    ("\n  altera 文件夹路径: ", "key"),
                    (f"{altera_dir}\n", "muted"),
                ])

            self._append_result_lines([
                ("\n" + "\u2500" * 50 + "\n", "line"),
            ])

            # ---- 热销关键词统计（来自 Title 列） ----
            if primary_result.get("hot_keywords") and primary_result["hot_keywords"].get("top_keywords"):
                hk = primary_result["hot_keywords"]
                self._append_result_lines([
                    ("\n  热销关键词统计（Title 列高频词）\n\n", "key"),
                    (f"  标题总数: {hk['total_titles']}  |  唯一关键词: {hk['total_unique_words']} 个  |  Top {len(hk['top_keywords'])}\n\n", "muted"),
                ])
                # 两列显示，更紧凑
                keywords = hk["top_keywords"]
                total = hk["total_titles"]
                half = (len(keywords) + 1) // 2
                for i in range(half):
                    left_rank = i + 1
                    left_kw, left_cnt = keywords[i]
                    left_pct = f"{left_cnt / total * 100:.1f}%" if total > 0 else "0%"
                    left_str = f"  {left_rank:>2}. {left_kw:<18} {left_cnt:>3}次 ({left_pct})"

                    right_idx = i + half
                    right_str = ""
                    if right_idx < len(keywords):
                        right_rank = right_idx + 1
                        right_kw, right_cnt = keywords[right_idx]
                        right_pct = f"{right_cnt / total * 100:.1f}%" if total > 0 else "0%"
                        right_str = f"    {right_rank:>2}. {right_kw:<18} {right_cnt:>3}次 ({right_pct})"

                    self._append_result_lines([
                        (left_str + right_str + "\n", "accent"),
                    ])

            # ---- 菜单层级树输出 ----
            menu_tree_lines = self._format_menu_tree(primary_result["menu_rows"])
            for line in menu_tree_lines:
                text, tag = line
                self._append_result_lines([(text, tag)])

            self._append_result_lines([
                ("\u2713 \u6587\u4ef6\u5df2\u751f\u6210\uff0c\u53ef\u7528\u4e8e Altera \u5bfc\u5165\u3002\n", "ok"),
            ])
            messagebox.showinfo("完成", "Excel 文件已生成")

        except Exception as e:
            import traceback
            self.result_text.configure(state="normal")
            self.result_text.delete("1.0", "end")
            self.result_text.configure(state="disabled")
            self._append_result_lines([
                ("\u2717 生成失败\n\n", "err"),
                (str(e) + "\n\n", "muted"),
                (traceback.format_exc() + "\n", "muted"),
            ])
            messagebox.showerror("生成失败", str(e))

    def on_open_file(self):
        if self._last_output_path and os.path.exists(self._last_output_path):
            try:
                if sys.platform.startswith("win"): os.startfile(self._last_output_path)
                elif sys.platform == "darwin": os.system(f'open "{self._last_output_path}"')
                else: os.system(f'xdg-open "{self._last_output_path}"')
            except Exception as e: messagebox.showerror("错误", str(e))
        else: messagebox.showinfo("提示", "请先生成文件")

    def on_open_folder(self):
        if self._last_output_path and os.path.exists(self._last_output_path):
            try:
                folder = os.path.dirname(os.path.abspath(self._last_output_path))
                if sys.platform.startswith("win"): os.startfile(folder)
                elif sys.platform == "darwin": os.system(f'open "{folder}"')
                else: os.system(f'xdg-open "{folder}"')
            except Exception as e: messagebox.showerror("错误", str(e))
        else: messagebox.showinfo("提示", "请先生成文件")

# ============================================================
# 启动入口
# ============================================================

if __name__ == "__main__":
    if len(sys.argv) >= 5 and sys.argv[1] == "--do-update":
        _run_update_worker(sys.argv[2], sys.argv[3], sys.argv[4])
        sys.exit(0)
    root = tk.Tk()
    app = AlteraExcelGeneratorApp(root)
    root.mainloop()
